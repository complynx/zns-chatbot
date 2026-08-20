import asyncio
import datetime
from motor.core import AgnosticCollection
from ..config import full_link
from ..tg_state import TGState
from telegram import InlineKeyboardMarkup, Update, InlineKeyboardButton, WebAppInfo
from .base_plugin import BasePlugin, PRIORITY_BASIC, PRIORITY_NOT_ACCEPTING
from telegram.ext import CommandHandler, CallbackQueryHandler, MessageHandler, filters
from telegram.constants import ParseMode
from bson.objectid import ObjectId
from pymongo.errors import DuplicateKeyError
from ..telegram_links import client_user_link_html, client_user_name
import logging
from .massage import now_msk
from math import ceil

def currency_ceil(sum):
    # if sum < 100:
    #     return ceil(sum)
    
    # # Get the magnitude (order of the largest digit) of the number
    # magnitude = 10 ** (len(str(int(sum))) - 2)
    
    # # Normalize the number by dividing by the magnitude
    # normalized = sum / magnitude
    
    # # Round up to the nearest 0 or 5
    # ceil_normalized = ceil(normalized * 2) / 2
    
    # # Scale back to the original magnitude
    # rounded = ceil_normalized * magnitude
    
    # return rounded
    return ceil(sum * 100) / 100

logger = logging.getLogger(__name__)

BYN_TO_RUB = 30
SHUTTLE_CAPACITY = 43
SHUTTLE_SERVICE = "shuttle"
GRODNO_OVERVIEW_SERVICE = "excursion_grodno_overview"
GRODNO_GORODNITSA_SERVICE = "excursion_grodno_gorodnitsa"
CAPACITY_LIMITS = {
    SHUTTLE_SERVICE: SHUTTLE_CAPACITY,
    GRODNO_OVERVIEW_SERVICE: 20,
    GRODNO_GORODNITSA_SERVICE: 25,
}

DEADLINE=datetime.datetime(2026, 9, 19, 0, 0, 0)


class CapacityFullError(Exception):
    def __init__(self, service):
        super().__init__(f"capacity is full for {service}")
        self.service = service


class ShuttleFullError(CapacityFullError):
    def __init__(self):
        super().__init__(SHUTTLE_SERVICE)


class InvalidExcursionChoiceError(ValueError):
    pass


def choice_has_shuttle(choice):
    extras = choice.get("extras", {}) if isinstance(choice, dict) else {}
    return isinstance(extras, dict) and SHUTTLE_SERVICE in extras


def choice_capacity_services(choice):
    extras = choice.get("extras", {}) if isinstance(choice, dict) else {}
    if not isinstance(extras, dict):
        return set()
    return {service for service in CAPACITY_LIMITS if service in extras}


def validate_excursion_choice(choice):
    extras = choice.get("extras", {}) if isinstance(choice, dict) else {}
    if not isinstance(extras, dict):
        raise InvalidExcursionChoiceError("extras must be an object")
    selected = {
        service for service in (GRODNO_OVERVIEW_SERVICE, GRODNO_GORODNITSA_SERVICE)
        if service in extras
    }
    if "excursion_grodno" in extras or len(selected) > 1:
        raise InvalidExcursionChoiceError("select exactly one Grodno excursion variant")


def capacity_full_error(service):
    if service == SHUTTLE_SERVICE:
        return ShuttleFullError()
    return CapacityFullError(service)

class OrdersUpdate:
    base: 'Orders'
    tgUpdate: Update
    user: int
    bot: int
    update: TGState
    _order = None

    def __init__(self, base, update: TGState) -> None:
        self.base = base
        self.update = update
        self.l = update.l
        self.user = update.user
        self.config = update.config.orders
        self.tgUpdate = update.update
        self.bot = self.update.bot.id

    async def create_order(self, choice):
        # Block creating orders after deadline
        if now_msk() > DEADLINE:
            return await self.update.reply(
                self.l("orders-closed"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )
        validate_excursion_choice(choice)
        order_id = ObjectId()
        reserved_services = []
        for service in sorted(choice_capacity_services(choice)):
            if not await self.base.reserve_service_seat(service, order_id):
                for reserved_service in reserved_services:
                    await self.base.release_service_seat(reserved_service, order_id)
                raise capacity_full_error(service)
            reserved_services.append(service)
        try:
            await self.base.food_db.insert_one({
                "_id": order_id,
                "user_id": self.user,
                "event_number": self.config.event_number,
                "created_at": datetime.datetime.now(),
                "choice": choice,
            })
        except Exception:
            for service in reserved_services:
                await self.base.release_service_seat(service, order_id)
            raise
        return await self.handle_cq_start()

    async def set_choice(self, order_id, choice):
        # Block modifying orders after deadline
        if now_msk() > DEADLINE:
            return await self.update.reply(
                self.l("orders-closed"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )
        validate_excursion_choice(choice)
        order_oid = ObjectId(order_id)
        previous_order = await self.base.food_db.find_one({"_id": order_oid})
        previous_services = choice_capacity_services(
            previous_order.get("choice", {}) if previous_order else {}
        )
        next_services = choice_capacity_services(choice)
        reserved_services = []
        for service in sorted(next_services - previous_services):
            if not await self.base.reserve_service_seat(service, order_oid):
                for reserved_service in reserved_services:
                    await self.base.release_service_seat(reserved_service, order_oid)
                raise capacity_full_error(service)
            reserved_services.append(service)
        try:
            result = await self.base.food_db.update_one({
                "_id": order_oid,
            },{
                "$set":{
                    "choice": choice,
                    "updated_at": datetime.datetime.now(),
                }
            })
            if result.matched_count == 0:
                raise ValueError(f"order {order_id} not found")
        except Exception:
            for service in reserved_services:
                await self.base.release_service_seat(service, order_oid)
            raise
        for service in previous_services - next_services:
            await self.base.release_service_seat(service, order_oid)
        return await self.handle_cq_start()

    async def handle_cq_del(self, order_id):
        # Disallow deleting after deadline
        if now_msk() > DEADLINE:
            return await self.handle_cq_start()
        order = await self.base.food_db.find_one({"_id": ObjectId(order_id)})
        if "proof_file" in order:
            return await self.handle_cq_start()
        await self.base.food_db.delete_one({"_id": ObjectId(order_id)})
        for service in choice_capacity_services(order.get("choice", {})):
            await self.base.release_service_seat(service, order["_id"])
        return await self.handle_cq_start()

    def get_order_total(self, order):
        total = currency_ceil(order["choice"]["total"])
        total_rub = currency_ceil(order["choice"]["total"] * BYN_TO_RUB)
        return total, total_rub
    
    async def handle_cq_pay(self, order_id):
        order = await self.base.food_db.find_one({"_id": ObjectId(order_id)})
        if "proof_file" in order:
            return await self.handle_cq_start()
        total, total_rub = self.get_order_total(order)
        admins_be = await self.base.base_app.users_collection.find({
            "bot_id": self.bot,
            "payment_administrator_belarus": {"$exists":True},
        }).to_list(None)
        btns = []
        for admin in admins_be:
            btns.append([
                InlineKeyboardButton(
                    self.l("orders-admin-belarus",
                        name=client_user_name(admin),
                        region=admin['payment_administrator_belarus']
                    ),
                    callback_data=f"{self.base.name}|cash|{order_id}|{admin['user_id']}"
                )
            ])
        await self.update.edit_or_reply(
            self.l("orders-message-payment-options",
                total=total,
                rutotal=total_rub,
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(btns+[[InlineKeyboardButton(
                self.l("orders-back-button"),
                callback_data=f"{self.base.name}|start"
            ), InlineKeyboardButton(
                self.l("orders-close-button"),
                callback_data=f"{self.base.name}|close"
            )]]),
        )

    async def handle_cq_cash(self, order_id, admin_id):
        # Block cash confirmation creation after deadline
        if now_msk() > DEADLINE:
            return await self.handle_cq_start()
        # return await self.handle_cq_start()
        admin = await self.base.base_app.users_collection.find_one({
            "user_id": int(admin_id),
            "bot_id": self.bot,
            "payment_administrator_belarus": {"$exists":True},
        })
        if admin is not None:
            await self.base.food_db.update_one({
                "_id": ObjectId(order_id),
            }, {
                "$set": {
                    "proof_country": "be",
                    "proof_admin": int(admin_id),
                    "proof_file": "cash",
                    "proof_received": datetime.datetime.now(),
                }
            })
            order = await self.base.food_db.find_one({"_id": ObjectId(order_id)})
            total, _total_rub = self.get_order_total(order)
            user = await self.update.get_user()
            lc = "ru"
            if "language_code" in admin:
                lc = admin["language_code"]
            def loc(s, **kwargs):
                return self.base.base_app.localization(s, args=kwargs, locale=lc)
            await self.update.reply(
                loc(
                    "orders-adm-payment-cash-requested",
                    link=client_user_link_html(user),
                    total=total,
                    name=order["choice"]["customer"],
                ),
                chat_id=admin["user_id"],
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton(loc("food-adm-payment-proof-accept-button"), callback_data=f"{self.base.name}|adm_acc|{order_id}"),
                        InlineKeyboardButton(loc("food-adm-payment-proof-reject-button"), callback_data=f"{self.base.name}|adm_rej|{order_id}"),
                    ]
                ])
            )
        await self.update.edit_or_reply(
            self.l("orders-payment-cash-requested",
                link=client_user_link_html(admin),
                total=total,
                name=order["choice"]["customer"],
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_cq_start(self):
        orders = await self.base.food_db.find({
            "user_id": self.user,
            "event_number": self.config.event_number,
        }).sort("created_at", 1).to_list(None)
        user = await self.update.get_user()
        debug_param = ""
        if "debug_id" in user:
            debug_param = "&debug_id="+user["debug_id"]
        current_order = None
        btns = []
        for order in orders:
            if "proof_file" not in order:
                current_order = order
                break
            else:
                btns.append([InlineKeyboardButton(self.l(
                    "orders-order-button",
                    created=order["created_at"].strftime("%d.%m"),
                    name=order["choice"]["customer"],
                ), web_app=WebAppInfo(full_link(self.base.base_app, f"/orders?order_id={str(order['_id'])}&locale={self.update.language_code}{debug_param}")))])
        if now_msk() <= DEADLINE:
            if current_order is not None:
                btns.append([InlineKeyboardButton(self.l(
                    "orders-order-pay-button",
                ), callback_data=f"{self.base.name}|pay|{str(order['_id'])}")])
                btns.append([InlineKeyboardButton(self.l(
                    "orders-order-unpaid-button",
                    created=order["created_at"].strftime("%d.%m"),
                    name=order["choice"]["customer"],
                ), web_app=WebAppInfo(full_link(self.base.base_app, f"/orders?order_id={str(order['_id'])}&locale={self.update.language_code}{debug_param}")))])
                btns.append([InlineKeyboardButton(
                    self.l("orders-edit-button"),
                    web_app=WebAppInfo(full_link(self.base.base_app, f"/orders?order_id={str(order['_id'])}&locale={self.update.language_code}{debug_param}"))
                )])
                btns.append([InlineKeyboardButton(self.l(
                    "orders-order-delete-button",
                ), callback_data=f"{self.base.name}|del|{str(order['_id'])}")])
            else:
                btns.append([InlineKeyboardButton(
                    self.l("orders-new-button"),
                    web_app=WebAppInfo(full_link(self.base.base_app, f"/orders?locale={self.update.language_code}{debug_param}"))
                )])
        # Admin download button
        admins_be = await self.base.base_app.users_collection.find({
            "bot_id": self.bot,
            "payment_administrator_belarus": {"$exists": True},
        }).to_list(None)
        admins_be_ids = {a["user_id"] for a in admins_be}
        if (self.user in self.config.admins or
            (self.config.payment_admin_ru and self.user == self.config.payment_admin_ru) or
            self.user in admins_be_ids):
            btns.append([InlineKeyboardButton("📥 XLSX", callback_data=f"{self.base.name}|xlsx")])
        btns.append([InlineKeyboardButton(
            self.l("orders-close-button"),
            callback_data=f"{self.base.name}|close"
        )])
        await self.update.edit_or_reply(self.l("orders-message-list"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(btns),
        )
    
    async def handle_cq_close(self):
        await self.update.edit_or_reply(self.l("orders-closed"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_callback_query(self):
        q = self.update.callback_query
        await q.answer()
        logger.info(f"Received callback_query from {self.user}, data: {q.data}")
        data = q.data.split("|")
        fn = "handle_cq_" + data[1]
        logger.debug(f"fn: {fn}")
        if hasattr(self, fn):
            attr = getattr(self, fn, None)
            logger.debug(f"fn: {attr}")
            if callable(attr):
                return await attr(*data[2:])
        logger.error(f"unknown callback {data[1]}: {data[2:]}")

    async def handle_start(self):
        logger.debug(f"starting orders for: {self.user}")
        await self.handle_cq_start()
    
    async def handle_cq_paid(self, order_id):
        # return await self.handle_cq_start()
        await self.base.food_db.update_one({
            "_id": ObjectId(order_id),
        }, {
            "$set": {
                "proof_country": "ru",
            }
        })
        order = await self.base.food_db.find_one({"_id": ObjectId(order_id)})
        _total_be, total = self.get_order_total(order)
        adm = self.base.config.orders.payment_admin_ru
        if adm>0:
            admin = await self.base.base_app.users_collection.find_one({
                "user_id": adm,
                "bot_id": self.bot,
            })
            user = await self.update.get_user()
            lc = "ru"
            if admin is not None and "language_code" in admin:
                lc = admin["language_code"]
            def loc(s, **kwargs):
                return self.base.base_app.localization(s, args=kwargs, locale=lc)
            await self.update.forward_message(
                adm,
                order["proof_chat_id"],
                order["proof_message_id"]
            )
            await self.update.reply(
                loc(
                    "orders-adm-payment-proof-received",
                    link=client_user_link_html(user),
                    total=total,
                    name=order["choice"]["customer"],
                ),
                chat_id=adm,
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton(loc("food-adm-payment-proof-accept-button"), callback_data=f"{self.base.name}|adm_acc|{order_id}"),
                        InlineKeyboardButton(loc("food-adm-payment-proof-reject-button"), callback_data=f"{self.base.name}|adm_rej|{order_id}"),
                    ]
                ])
            )
        await self.update.edit_or_reply(
            self.l("food-payment-proof-forwarded"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_cq_adm_acc(self, order_id):
        # Guard: only RU / BE payment admins or orders admins (assert style)
        admins_be_ids = {a["user_id"] for a in await self.base.base_app.users_collection.find({
            "bot_id": self.bot,
            "payment_administrator_belarus": {"$exists": True},
        }).to_list(None)}
        assert (self.user in self.config.admins or
                (self.config.payment_admin_ru and self.user == self.config.payment_admin_ru) or
                self.user in admins_be_ids), f"{self.user} is not orders admin"
        await self.base.food_db.update_one({
            "_id": ObjectId(order_id)
        },{
            "$set":{
                "validated_at": datetime.datetime.now(),
                "validation": True,
            }
        })
        order = await self.base.food_db.find_one({"_id": ObjectId(order_id)})
        user = await self.base.base_app.users_collection.find_one({
            "user_id": order["user_id"],
            "bot_id": self.bot,
        })
        ls = 'en'
        if "language_code" in user:
            ls=user["language_code"]
        def loc(s, **kwargs):
            return self.base.base_app.localization(s, args=kwargs, locale=ls)
        await self.update.reply(
            loc(
                "food-payment-proof-confirmed",
                name=order["choice"]["customer"],
            ),
            order["user_id"],
            parse_mode=ParseMode.HTML
        )
        await self.update.edit_message_text(
            self.l(
                "food-adm-payment-proof-confirmed",
                link=client_user_link_html(user),
                name=order["choice"]["customer"],
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([])
        )

    async def handle_cq_adm_rej(self, order_id):
        # Guard: only RU / BE payment admins or orders admins (assert style)
        admins_be_ids = {a["user_id"] for a in await self.base.base_app.users_collection.find({
            "bot_id": self.bot,
            "payment_administrator_belarus": {"$exists": True},
        }).to_list(None)}
        assert (self.user in self.config.admins or
                (self.config.payment_admin_ru and self.user == self.config.payment_admin_ru) or
                self.user in admins_be_ids), f"{self.user} is not orders admin"
        await self.base.food_db.update_one({
            "_id": ObjectId(order_id)
        },{
            "$set":{
                "validated_at": datetime.datetime.now(),
                "validation": False,
            },
            "$unset": {
                "proof_file": "",
                "proof_received": "",
                "proof_chat_id": "",
                "proof_message_id": "",
            },
        })
        order = await self.base.food_db.find_one({"_id": ObjectId(order_id)})
        user = await self.base.base_app.users_collection.find_one({
            "user_id": order["user_id"],
            "bot_id": self.bot,
        })
        ls = 'en'
        if "language_code" in user:
            ls=user["language_code"]
        def loc(s, **kwargs):
            return self.base.base_app.localization(s, args=kwargs, locale=ls)
        await self.update.reply(
            loc(
                "food-payment-proof-rejected",
                name=order["choice"]["customer"],
            ),
            order["user_id"],
            parse_mode=ParseMode.HTML
        )
        await self.update.edit_message_text(
            self.l(
                "food-adm-payment-proof-rejected",
                link=client_user_link_html(user),
                name=order["choice"]["customer"],
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([])
        )
    
    async def handle_cq_pcancel(self, order_id):
        # Block canceling proof after deadline
        if now_msk() > DEADLINE:
            return await self.update.edit_or_reply(self.l("orders-closed"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )
        await self.base.food_db.update_one({"_id": ObjectId(order_id)}, {
            "$unset": {
                "proof_file": "",
                "proof_received": "",
                "proof_chat_id": "",
                "proof_message_id": "",
            }
        })
        await self.update.edit_or_reply(self.l("orders-closed"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_payment(self):
        logger.debug(f"handling payment for: {self.user}")
        # After deadline, ignore new payment proofs
        if now_msk() > DEADLINE:
            return await self.update.edit_or_reply(
                self.l("orders-closed"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )
        order = await self.base.food_db.find_one({
            "user_id": self.user,
            "event_number": self.config.event_number,
            "proof_file": { "$exists": False },
        })
        if order is None:
            return await self.update.edit_or_reply(
                self.l("unsupported-message-error"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )
        doc = self.update.message.document
        await self.base.food_db.update_one({
            "_id": order["_id"]
        }, {
            "$set": {
                "proof_file": doc.file_id,
                "proof_chat_id": self.update.chat_id if self.update.chat_id is not None else self.update.user,
                "proof_message_id": self.update.message_id,
                "proof_received": datetime.datetime.now(),
            }
        })
        await self.update.edit_or_reply(
            self.l("orders-message-paid-where"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(
                self.l("orders-paid-button"),
                callback_data=f"{self.base.name}|paid|{order['_id']}"
            )],[InlineKeyboardButton(
                self.l("orders-pay-cancel"),
                callback_data=f"{self.base.name}|pcancel|{order['_id']}"
            )]]),
        )

    async def handle_cq_xlsx(self):
        # Guard (assert style like passes)
        admins_be = await self.base.base_app.users_collection.find({
            "bot_id": self.bot,
            "payment_administrator_belarus": {"$exists": True},
        }).to_list(None)
        admins_be_ids = {a["user_id"] for a in admins_be}
        assert (self.user in self.config.admins or
                (self.config.payment_admin_ru and self.user == self.config.payment_admin_ru) or
                self.user in admins_be_ids), f"{self.user} is not orders admin"
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказы"
        # Collect menu and disposable-service columns.
        menu_dishes = self.base.menu.get("dishes", {})
        dish_keys = list(menu_dishes)
        dish_names_ru = {
            dish_key: dish.get("name_ru", dish_key)
            for dish_key, dish in menu_dishes.items()
        }
        service_items = self.base.menu.get("service_items", {})
        service_keys = list(service_items)
        service_names_ru = {
            service_key: service.get("name_ru", service_key)
            for service_key, service in service_items.items()
        }
        # Base columns (english keys -> russian headers)
        base_fields = [
            ("order_id", "ID заказа"),
            ("user_id", "Пользователь"),
            ("customer", "Клиент"),
            ("created_at", "Создан"),
            ("updated_at", "Обновлён"),
            ("proof_country", "Страна оплаты"),
            ("proof_admin", "Администратор оплаты"),
            ("validation", "Подтверждено"),
            ("total_byn", "Сумма BYN"),
            ("total_rub", "Сумма RUB"),
            ("extras_preparty", "Препати"),
            ("extras_excursion_minsk", "Экскурсия Минск"),
            ("extras_shuttle", "Трансфер"),
            ("extras_excursion_grodno_overview", "Гродно: обзорная"),
            ("extras_excursion_grodno_gorodnitsa", "Гродно: Городница"),
            ("extras_excursion_grodno", "Гродно: вариант не указан"),
        ]
        header = (
            [ru for _k, ru in base_fields]
            + [dish_names_ru.get(k, k) for k in dish_keys]
            + [service_names_ru.get(k, k) for k in service_keys]
        )
        ws.append(header)
        bold = Font(bold=True)
        center = Alignment(horizontal="center")
        for cell in ws["1:1"]:
            cell.font = bold
            cell.alignment = center
        totals = {k: {"count":0,"sum":0} for k in dish_keys}
        service_totals = {k: {"count":0,"sum":0} for k in service_keys}
        extras_totals = {
            "preparty": 0,
            "excursion_minsk": 0,
            "shuttle": 0,
            GRODNO_OVERVIEW_SERVICE: 0,
            GRODNO_GORODNITSA_SERVICE: 0,
            "excursion_grodno": 0,
        }
        # Second sheet with detailed contents
        ws_details = wb.create_sheet("Содержимое")
        ws_details.append(["Пользователь","ID заказа","Клиент","День","Приём пищи","Блюдо / Активность","Оплата","Количество"])
        for cell in ws_details["1:1"]:
            cell.font = bold
            cell.alignment = center
        day_ru = {"friday":"Пятница","saturday":"Суббота","sunday":"Воскресенье"}
        meal_ru = {"lunch":"Обед","dinner":"Ужин"}
        extras_ru = {
            "preparty":"Препати",
            "excursion_minsk":"Экскурсия по Минску",
            "shuttle":"Трансфер Минск–Гродно",
            GRODNO_OVERVIEW_SERVICE: "Гродно: обзорная экскурсия",
            GRODNO_GORODNITSA_SERVICE: "Гродно: экскурсия «Городница»",
            "excursion_grodno":"Экскурсия по Гродно (вариант не указан)",
        }
        async for order in self.base.food_db.find({"event_number": self.config.event_number}):
            choice = order.get("choice", {})
            total_byn = choice.get("total", 0)
            total_rub = total_byn * BYN_TO_RUB
            dish_counts = {k:0 for k in dish_keys}
            service_counts = {k:0 for k in service_keys}
            # Detailed dishes
            for day_key, day_data in choice.get("days", {}).items():
                for mealtime_key, mealtime in day_data.get("mealtimes", {}).items():
                    for dish in mealtime.get("dishes", []):
                        name_key = dish.get("name")
                        cnt = dish.get("count",0)
                        price = dish.get("price",0)
                        ru_name = dish_names_ru.get(name_key, name_key)
                        ws_details.append([
                            order.get("user_id",""),
                            str(order.get("_id")),
                            choice.get("customer",""),
                            day_ru.get(day_key, day_key),
                            meal_ru.get(mealtime_key, mealtime_key),
                            ru_name,
                            order.get("validation",""),
                            cnt,
                        ])
                        if name_key in dish_counts:
                            dish_counts[name_key] += cnt
                            totals[name_key]["count"] += cnt
                            totals[name_key]["sum"] += cnt*price
                    for service in mealtime.get("service", {}).get("items", []):
                        name_key = service.get("name")
                        cnt = service.get("count", 0)
                        price = service.get("price", 0)
                        ws_details.append([
                            order.get("user_id", ""),
                            str(order.get("_id")),
                            choice.get("customer", ""),
                            day_ru.get(day_key, day_key),
                            meal_ru.get(mealtime_key, mealtime_key),
                            service_names_ru.get(name_key, name_key),
                            order.get("validation", ""),
                            cnt,
                        ])
                        if name_key in service_counts:
                            service_counts[name_key] += cnt
                            service_totals[name_key]["count"] += cnt
                            service_totals[name_key]["sum"] += cnt * price
            # Extras rows
            extras = choice.get("extras", {})
            for ex_key, ex_ru in extras_ru.items():
                if ex_key in extras:
                    ws_details.append([
                        order.get("user_id",""),
                        str(order.get("_id")),
                        choice.get("customer",""),
                        day_ru.get("friday","Пятница"),  # day not specified -> reuse first day label or blank
                        "активности",
                        ex_ru,
                        order.get("validation",""),
                        1,
                    ])
                    extras_totals[ex_key]+=1
            row = [
                str(order.get("_id")),
                order.get("user_id",""),
                choice.get("customer",""),
                order.get("created_at",""),
                order.get("updated_at",""),
                order.get("proof_country",""),
                order.get("proof_admin",""),
                order.get("validation",""),
                currency_ceil(total_byn),
                currency_ceil(total_rub),
                1 if "preparty" in extras else 0,
                1 if "excursion_minsk" in extras else 0,
                1 if "shuttle" in extras else 0,
                1 if GRODNO_OVERVIEW_SERVICE in extras else 0,
                1 if GRODNO_GORODNITSA_SERVICE in extras else 0,
                1 if "excursion_grodno" in extras else 0,
            ] + [dish_counts[k] for k in dish_keys] + [service_counts[k] for k in service_keys]
            ws.append(row)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is None:
                        continue
                    length = len(str(cell.value))
                    if length>max_length:
                        max_length=length
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_length, 6), 40)
        for col in ws_details.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is None:
                        continue
                    length = len(str(cell.value))
                    if length>max_length:
                        max_length=length
                except Exception:
                    pass
            ws_details.column_dimensions[col_letter].width = min(max(max_length, 6), 40)
        ws_totals = wb.create_sheet("Итоги")
        ws_totals.append(["Блюдо","Количество","Сумма BYN"])
        for cell in ws_totals["1:1"]:
            cell.font = bold
            cell.alignment = center
        for k,v in totals.items():
            ws_totals.append([dish_names_ru.get(k,k), v["count"], currency_ceil(v["sum"])])
        ws_service = wb.create_sheet("Посуда")
        ws_service.append(["Позиция", "Количество", "Сумма BYN"])
        for cell in ws_service["1:1"]:
            cell.font = bold
            cell.alignment = center
        for k, v in service_totals.items():
            ws_service.append([
                service_names_ru.get(k, k),
                v["count"],
                currency_ceil(v["sum"]),
            ])
        ws_extras = wb.create_sheet("Активности")
        ws_extras.append(["Активность","Кол-во заказов"])
        for cell in ws_extras["1:1"]:
            cell.font = bold
            cell.alignment = center
        for k,v in extras_totals.items():
            ws_extras.append([extras_ru.get(k,k), v])
        file_name = "orders.xlsx"
        wb.save(file_name)
        await self.update.bot.send_document(
            self.update.user,
            open(file_name, "rb"),
            caption="Orders XLSX",
        )
        import os
        os.remove(file_name)
        await self.handle_cq_start()

class Orders(BasePlugin):
    name = "orders"
    food_db: AgnosticCollection

    def __init__(self, base_app):
        super().__init__(base_app)
        self.base_app.orders = self
        self.food_db = base_app.mongodb[self.config.mongo_db.food_collection]
        self.capacity_db = base_app.mongodb[
            self.config.mongo_db.food_collection + "_capacity"
        ]
        self._capacity_slots_ready = set()
        self._capacity_slots_lock = asyncio.Lock()
        self._checker = CommandHandler(self.name, self.handle_start)
        self._file_checker = MessageHandler(filters.Document.PDF, self.handle_payment)
        self._cbq_handler = CallbackQueryHandler(self.handle_callback_query, pattern=f"^{self.name}\\|.*")
        self.menu = self.get_menu()

    def _capacity_event_number(self):
        return self.config.event_number

    async def _ensure_capacity_slots(self, service):
        capacity = CAPACITY_LIMITS[service]
        event_number = self._capacity_event_number()
        ready_key = (event_number, service)
        if ready_key in self._capacity_slots_ready:
            return
        async with self._capacity_slots_lock:
            if ready_key in self._capacity_slots_ready:
                return

            await self.capacity_db.create_index(
                [
                    ("event_number", 1),
                    ("service", 1),
                    ("reservation_id", 1),
                ],
                unique=True,
                partialFilterExpression={"reservation_id": {"$exists": True}},
            )
            for seat in range(capacity):
                await self.capacity_db.update_one(
                    {"_id": f"{event_number}:{service}:{seat}"},
                    {"$setOnInsert": {
                        "event_number": event_number,
                        "service": service,
                        "seat": seat,
                    }},
                    upsert=True,
                )

            existing_orders = await self.food_db.find(
                {
                    "event_number": event_number,
                    f"choice.extras.{service}": {"$exists": True},
                },
                {"_id": 1},
            ).to_list(None)
            existing_order_ids = {order["_id"] for order in existing_orders}
            reserved_slots = await self.capacity_db.find(
                {
                    "event_number": event_number,
                    "service": service,
                    "reservation_id": {"$exists": True},
                },
                {"reservation_id": 1},
            ).to_list(None)
            reserved_order_ids = {
                slot["reservation_id"] for slot in reserved_slots
                if slot.get("reservation_id") in existing_order_ids
            }
            for slot in reserved_slots:
                reservation_id = slot.get("reservation_id")
                if reservation_id not in existing_order_ids:
                    await self.capacity_db.update_one(
                        {
                            "_id": slot["_id"],
                            "reservation_id": reservation_id,
                        },
                        {"$unset": {"reservation_id": "", "reserved_at": ""}},
                    )

            for order_id in existing_order_ids - reserved_order_ids:
                await self.capacity_db.find_one_and_update(
                    {
                        "event_number": event_number,
                        "service": service,
                        "reservation_id": {"$exists": False},
                    },
                    {"$set": {
                        "reservation_id": order_id,
                        "reserved_at": datetime.datetime.now(),
                    }},
                    sort=[("seat", 1)],
                )

            self._capacity_slots_ready.add(ready_key)

    async def reserve_service_seat(self, service, order_id):
        await self._ensure_capacity_slots(service)
        event_number = self._capacity_event_number()
        existing = await self.capacity_db.find_one({
            "event_number": event_number,
            "service": service,
            "reservation_id": order_id,
        })
        if existing is not None:
            return True
        try:
            claimed = await self.capacity_db.find_one_and_update(
                {
                    "event_number": event_number,
                    "service": service,
                    "reservation_id": {"$exists": False},
                },
                {"$set": {
                    "reservation_id": order_id,
                    "reserved_at": datetime.datetime.now(),
                }},
                sort=[("seat", 1)],
            )
        except DuplicateKeyError:
            # A concurrent request may have reserved another slot for this order.
            claimed = await self.capacity_db.find_one({
                "event_number": event_number,
                "service": service,
                "reservation_id": order_id,
            })
        return claimed is not None

    async def release_service_seat(self, service, order_id):
        await self._ensure_capacity_slots(service)
        await self.capacity_db.update_one(
            {
                "event_number": self._capacity_event_number(),
                "service": service,
                "reservation_id": order_id,
            },
            {"$unset": {"reservation_id": "", "reserved_at": ""}},
        )

    async def service_available(self, service, current_choice=None):
        if service in choice_capacity_services(current_choice or {}):
            return True
        await self._ensure_capacity_slots(service)
        free_slot = await self.capacity_db.find_one({
            "event_number": self._capacity_event_number(),
            "service": service,
            "reservation_id": {"$exists": False},
        })
        return free_slot is not None

    async def reserve_shuttle_seat(self, order_id):
        return await self.reserve_service_seat(SHUTTLE_SERVICE, order_id)

    async def release_shuttle_seat(self, order_id):
        await self.release_service_seat(SHUTTLE_SERVICE, order_id)

    async def shuttle_available(self, current_choice=None):
        return await self.service_available(SHUTTLE_SERVICE, current_choice)

    def get_menu(self):
        from os.path import dirname as d
        from os.path import join
        import json
        menu_file = join(d(d(d(__file__))), "static", "menu_belarus.json")
        with open(menu_file, "r", encoding="utf-8") as mf:
            return json.load(mf)

    async def create_update_from_user(self, user) -> OrdersUpdate:
        upd = TGState(user, self.base_app)
        await upd.get_state()
        return OrdersUpdate(self, upd)
    
    async def create_order(self, user_id, choice):
        upd = await self.create_update_from_user(user_id)
        return await upd.create_order(choice)
    
    async def set_choice(self, order, choice):
        upd = await self.create_update_from_user(order["user_id"])
        return await upd.set_choice(order["_id"], choice)

    async def order_by_id(self, order_id):
        return await self.food_db.find_one({"_id": ObjectId(order_id)})

    def test_message(self, message: Update, state, web_app_data):
        if self._checker.check_update(message):
            return PRIORITY_BASIC, self.handle_start
        if self._file_checker.check_update(message):
            return PRIORITY_BASIC, self.handle_payment
        return PRIORITY_NOT_ACCEPTING, None
    
    def test_callback_query(self, query: Update, state):
        if self._cbq_handler.check_update(query):
            return PRIORITY_BASIC, self.handle_callback_query
        return PRIORITY_NOT_ACCEPTING, None
    
    def create_update(self, update) -> OrdersUpdate:
        return OrdersUpdate(self, update)
    
    async def handle_callback_query(self, updater):
        return await self.create_update(updater).handle_callback_query()

    async def handle_start(self, update: TGState):
        return await self.create_update(update).handle_start()

    async def handle_payment(self, update: TGState):
        return await self.create_update(update).handle_payment()
