from asyncio import Event, Lock, create_task, sleep
from datetime import timedelta
from random import choice
import logging

from motor.core import AgnosticCollection
from telegram import (
    Contact,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    MessageOriginUser,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.constants import ParseMode
from telegram.ext import CallbackQueryHandler, CommandHandler, filters

from ..events import EventInfo, EventPassType, Events
from ..telegram_links import client_user_link_html, client_user_name
from ..tg_state import SilentArgumentParser, TGState
from .base_plugin import PRIORITY_BASIC, PRIORITY_NOT_ACCEPTING, BasePlugin
from .massage import now_msk, split_list

logger = logging.getLogger(__name__)

CANCEL_CHR = chr(0xE007F)  # Tag cancel
MAX_CONCURRENT_ASSIGNMENTS = 6
QUEUE_LOOKAHEAD = 0  # 0 = scan the whole queue when searching for an assignable pass

# Per-person default pass prices by pass key.
CURRENT_PRICE = {
    "pass_2026_1": 12500,
}
TIMEOUT_PROCESSOR_TICK = 3600
INVITATION_TIMEOUT = timedelta(days=2, hours=10)
PAYMENT_TIMEOUT = timedelta(days=8)
PAYMENT_TIMEOUT_NOTIFY2 = timedelta(days=7)
PAYMENT_TIMEOUT_NOTIFY = timedelta(days=6)
PASS_TYPES_ASSIGNABLE = ["solo", "couple", "sputnik"]
ROLE_BALANCE_TOLERANCE = 52  # percent of higher_role/total


class PassUpdate:
    base: "Passes"
    tgUpdate: Update
    bot: int
    update: TGState

    def __init__(self, base, update: TGState) -> None:
        self.base = base
        self.update = update
        self.l = update.l
        self.tgUpdate = update.update
        self.bot = self.update.bot.id
        self.pass_key = ""
        self.user = None
        self.pass_owner_id: int | None = None
        self.pass_data: dict | None = None

    def _pass_localization_keys(
        self,
        pass_key: str | None = None,
        locale: str | None = None,
    ) -> dict[str, str]:
        key = pass_key if pass_key is not None else self.pass_key
        if key == "":
            return {}
        locale_code = locale if locale is not None else self.update.language_code
        return self.base.get_event_localization_keys(key, locale_code)

    def l_pass(
        self,
        message_id: str,
        pass_key: str | None = None,
        locale: str | None = None,
        **kwargs,
    ) -> str:
        key = pass_key if pass_key is not None else self.pass_key
        if key:
            kwargs.setdefault("passKey", key)
            for loc_key, loc_value in self._pass_localization_keys(
                pass_key=key,
                locale=locale,
            ).items():
                kwargs.setdefault(loc_key, loc_value)
        return self.l(message_id, **kwargs)

    def is_passport_required(self) -> bool:
        event = self.base.get_event(self.pass_key)
        return event.require_passport if event is not None else True

    async def handle_callback_query(self):
        q = self.update.callback_query
        await q.answer()
        logger.info(f"Received callback_query from {self.update.user}, data: {q.data}")
        data = q.data.split("|")
        fn = "handle_cq_" + data[1]
        logger.debug(f"fn: {fn}")
        if hasattr(self, fn):
            attr = getattr(self, fn, None)
            logger.debug(f"fn: {attr}")
            if callable(attr):
                return await attr(*data[2:])
        logger.error(f"unknown callback {data[1]}: {data[2:]}")

    def admin_to_keys(self, admin: dict, lc: str | None = None) -> dict:
        assert isinstance(admin, dict)
        if lc is None:
            lc = self.update.language_code
        keys = {}
        keys["adminEmoji"] = admin.get("emoji", "")
        keys["adminLink"] = client_user_link_html(admin, language_code=lc)
        keys["adminName"] = client_user_name(admin, language_code=lc)
        keys["phoneContact"] = admin.get("phone_contact", "nophone")
        keys["phoneSBP"] = admin.get("phone_sbp", "nosbp")
        if admin.get("paypal", "") != "" and keys["phoneSBP"] == "nosbp":
            keys["phoneSBP"] = "paypal"
        keys["banks"] = admin.get("banks_en", "")
        if lc is not None:
            keys["banks"] = admin.get("banks_" + lc, keys["banks"])
        return keys

    async def format_message(
        self,
        message_id: str,
        user: int | dict | None = None,
        for_user: int | None = None,
        u_pass: dict | None = None,
        couple: int | dict | None = None,
        admin: int | dict | None = None,
        pass_key: str | None = None,
        **additional_keys,
    ) -> str:
        if pass_key is None:
            pass_key = self.pass_key
        if user is None:
            user = self.update.user
        if not isinstance(user, dict):
            user = await self.base.user_db.find_one(
                {
                    "user_id": user,
                    "bot_id": self.bot,
                }
            )
        if not isinstance(user, dict):
            logger.error(
                f"format_message.user is not a dict: {user=}, {type(user)=}",
                stack_info=True,
            )
            return message_id
        l = self.l  # noqa: E741
        lc = self.update.language_code
        if for_user is not None:
            upd = await self.base.create_update_from_user(for_user)
            l = upd.l  # noqa: E741
            lc = upd.update.language_code
        if not isinstance(u_pass, dict):
            u_pass = await self.get_pass(pass_key, user.get("user_id"))
        if not isinstance(u_pass, dict):
            u_pass = {}
        keys = dict(u_pass)
        couple_pass = None
        keys["passKey"] = pass_key
        keys.update(self.base.get_event_localization_keys(pass_key, lc))
        keys["name"] = user.get("legal_name", "")
        keys["link"] = client_user_link_html(user, language_code=lc)
        if couple is None:
            couple = u_pass.get("couple")
        if isinstance(couple, int):
            couple = await self.base.user_db.find_one(
                {
                    "user_id": couple,
                    "bot_id": self.bot,
                }
            )
        if isinstance(couple, dict):
            couple_pass = await self.get_pass(
                pass_key, couple.get("user_id")
            )
            keys["coupleName"] = couple.get("legal_name")
            keys["coupleRole"] = couple_pass.get("role") if isinstance(couple_pass, dict) else ""
            keys["coupleLink"] = client_user_link_html(couple, language_code=lc)
            if "type" not in keys:
                keys["type"] = "couple"
        elif "type" not in keys:
            keys["type"] = "solo"
        keys["price"] = self.base.get_pass_display_price(u_pass, couple_pass)
        if admin is None:
            if "proof_admin" in u_pass:
                admin = u_pass["proof_admin"]
        if isinstance(admin, int):
            admin = await self.base.user_db.find_one(
                {
                    "user_id": admin,
                    "bot_id": self.bot,
                }
            )
        if isinstance(admin, dict):
            keys.update(self.admin_to_keys(admin, lc))
        keys.update(additional_keys)
        return l(message_id, **keys)

    async def show_pass_edit(self, user, u_pass, text_key: str | None = None):
        if text_key is None:
            text_key = f"passes-pass-edit-{u_pass['state']}"
        state = u_pass.get("state")
        pass_type = u_pass.get("type", "solo")

        buttons = [
            InlineKeyboardButton(
                self.l("passes-button-change-name"),
                callback_data=f"{self.base.name}|name",
            ),
        ]
        if state == "assigned":
            buttons.append(
                InlineKeyboardButton(
                    self.l("passes-button-pay"),
                    callback_data=f"{self.base.name}|pay|{self.pass_key}",
                )
            )
        if state == "waitlist" and pass_type == "solo" and "couple" not in u_pass:
            buttons.append(
                InlineKeyboardButton(
                    self.l("passes-button-make-couple"),
                    callback_data=f"{self.base.name}|couple|{self.pass_key}",
                )
            )
        if state == "waiting-for-couple":
            buttons.append(
                InlineKeyboardButton(
                    self.l("passes-button-make-solo"),
                    callback_data=f"{self.base.name}|solo|{self.pass_key}",
                )
            )
        if state != "paid":
            if len(self.base.payment_admins[self.pass_key]) > 1:
                buttons.append(
                    InlineKeyboardButton(
                        self.l("passes-button-change-admin"),
                        callback_data=f"{self.base.name}|change_admin|{self.pass_key}",
                    )
                )
            buttons.append(
                InlineKeyboardButton(
                    self.l("passes-button-cancel"),
                    callback_data=f"{self.base.name}|cancel|{self.pass_key}",
                )
            )
        buttons.append(
            InlineKeyboardButton(
                self.l("passes-button-exit"), callback_data=f"{self.base.name}|exit"
            )
        )

        await self.update.edit_or_reply(
            await self.format_message(
                text_key,
                user=user,
                u_pass=u_pass,
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(split_list(buttons, 2)),
        )

    async def handle_cq_change_admin(
        self, key, message_key: str = "passes-choose-admin"
    ):
        self.set_pass_key(key)
        payment_admins = await self.base.user_db.find(
            {
                "bot_id": self.bot,
                "user_id": {"$in": self.base.payment_admins[self.pass_key]},
            }
        ).to_list(None)
        btns = []
        admin_texts = []
        for admin in payment_admins:
            admin_keys = self.admin_to_keys(admin)
            btns.append(
                InlineKeyboardButton(
                    self.l(
                        "passes-payment-admin-button",
                        **admin_keys,
                    ),
                    callback_data=f"{self.base.name}|cha2|{self.pass_key}|{admin['user_id']}",
                )
            )
            admin_texts.append(
                self.l(
                    "passes-payment-admin-desc",
                    **admin_keys,
                )
            )
        btns.append(
            InlineKeyboardButton(
                self.l("passes-button-exit"),
                callback_data=f"{self.base.name}|exit",
            )
        )
        await self.update.edit_or_reply(
            self.l(
                message_key,
                adminTexts="\n".join(
                    [f"{i + 1}. {t}" for i, t in enumerate(admin_texts)]
                ),
            ),
            reply_markup=InlineKeyboardMarkup(split_list(btns, 2)),
            parse_mode=ParseMode.HTML,
        )

    async def handle_cq_cha2(self, key, admin_id_str):
        self.set_pass_key(key)
        admin_id = int(admin_id_str)
        assert admin_id in self.base.payment_admins[self.pass_key]
        pass_data = await self.get_pass(key)
        if pass_data is not None:  # has pass
            pass_data = dict(pass_data)
            pass_data["proof_admin"] = admin_id
            await self.base.save_pass_data(self.update.user, self.pass_key, pass_data)
            user, _ = await self.base.get_user_with_pass(self.update.user, self.pass_key)
            await self.show_pass_edit(user or {}, pass_data)
        else:  # no pass yet
            updated = await self.base.user_db.update_one(
                {
                    "user_id": self.update.user,
                    "bot_id": self.update.bot.id,
                },
                {
                    "$set": {
                        "proof_admins." + self.pass_key: admin_id,
                    }
                },
            )
            assert updated.matched_count > 0
            await self.update.edit_or_reply(
                self.l(
                    "passes-pass-admin-saved",
                ),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                self.l("passes-button-solo"),
                                callback_data=f"{self.base.name}|solo|{self.pass_key}",
                            ),
                            InlineKeyboardButton(
                                self.l("passes-button-couple"),
                                callback_data=f"{self.base.name}|couple|{self.pass_key}",
                            ),
                        ],
                        [
                            InlineKeyboardButton(
                                self.l("passes-button-cancel"),
                                callback_data=f"{self.base.name}|pass_exit",
                            ),
                        ],
                    ]
                ),
            )

    async def handle_cq_pass_exit(self):
        await self.update.edit_or_reply(
            self.l("passes-pass-create-cancel"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_cq_exit(self):
        await self.update.edit_or_reply(
            self.l("passes-pass-exit"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_cq_cancel(self, key):
        self.set_pass_key(key)
        success = await self.cancel_pass()
        if success:
            await self.update.edit_or_reply(
                self.l("passes-pass-cancelled"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )
            await self.base.recalculate_queues()
        else:
            await self.update.edit_or_reply(
                self.l("passes-pass-cancel-failed"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )

    async def show_couple_invitation(self, inviter) -> None:
        await self.update.reply(
            await self.format_message(
                "passes-couple-invitation",
                couple=inviter,
            ),
            reply_markup=InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(
                            self.l("passes-button-couple-decline"),
                            callback_data=f"{self.base.name}|cp_decl|{self.pass_key}|{inviter['user_id']}",
                        ),
                        InlineKeyboardButton(
                            self.l("passes-button-couple-accept"),
                            callback_data=f"{self.base.name}|cp_acc|{self.pass_key}|{inviter['user_id']}",
                        ),
                    ]
                ]
            ),
            parse_mode=ParseMode.HTML,
        )

    async def handle_cq_cp_acc(self, key, inviter_id):
        self.set_pass_key(key)
        inviter_id = int(inviter_id)
        await self.update.edit_message_text(
            self.l("passes-accept-pass-request-name"),
            reply_markup=InlineKeyboardMarkup([]),
            parse_mode=ParseMode.HTML,
        )
        if not self.is_passport_required():
            return await self.after_legal_name_input(inviter_id, "skip")
        await self.handle_name_request(inviter_id)

    async def accept_invitation(self, inviter_id):
        inviter_id = int(inviter_id)
        user = await self.get_user()
        inviter, inviter_pass = await self.base.get_user_with_pass(inviter_id, self.pass_key)
        if inviter is not None and isinstance(inviter_pass, dict):
            inviter_pass = dict(inviter_pass)
            inviter_pass["state"] = "waitlist"
            inviter_pass["couple"] = self.update.user
            await self.base.save_pass_data(inviter_id, self.pass_key, inviter_pass)
            u_pass = dict(inviter_pass)
            u_pass["role"] = "leader" if u_pass["role"].startswith("f") else "follower"
            u_pass["couple"] = inviter_id
            await self.base.save_pass_data(self.update.user, self.pass_key, u_pass)
            inv_update = await self.base.create_update_from_user(inviter_id)
            inv_update.set_pass_key(self.pass_key)
            await inv_update.handle_invitation_accepted(user)
            await self.update.reply(
                self.l_pass("passes-invitation-successfully-accepted"),
                parse_mode=ParseMode.HTML,
            )

            return await self.base.recalculate_queues()
        await self.update.reply(
            self.l_pass("passes-invitation-accept-failed"),
            parse_mode=ParseMode.HTML,
        )

    async def handle_cq_cp_decl(self, key, inviter_id):
        self.set_pass_key(key)
        inviter_id = int(inviter_id)
        user = await self.get_user()
        inviter_pass = await self.base.get_pass_for_user(inviter_id, self.pass_key)
        if inviter_pass is not None and inviter_pass.get("couple") == self.update.user:
            inv_update = await self.base.create_update_from_user(inviter_id)
            inv_update.set_pass_key(key)
            await inv_update.handle_invitation_declined(user)
        await self.update.edit_or_reply(
            self.l_pass("passes-invitation-successfully-declined"),
            reply_markup=InlineKeyboardMarkup([]),
            parse_mode=ParseMode.HTML,
        )

    async def handle_invitation_accepted(self, invitee):
        await self.update.edit_or_reply(
            await self.format_message("passes-invitation-was-accepted", couple=invitee),
            parse_mode=ParseMode.HTML,
        )

    async def handle_invitation_declined(self, invitee):
        pass_data = await self.get_pass()
        success = False
        if isinstance(pass_data, dict):
            success = pass_data.get("couple") == invitee["user_id"]
            if success:
                updated_pass = dict(pass_data)
                updated_pass.pop("couple", None)
                await self.base.save_pass_data(self.update.user, self.pass_key, updated_pass)
        else:
            result = await self.base.pass_db.update_one(
                {
                    "bot_id": self.bot,
                    "user_id": self.update.user,
                    "pass_key": self.pass_key,
                    "couple": invitee["user_id"],
                },
                {"$unset": {"couple": ""}},
            )
            if result.modified_count > 0:
                success = True

        if success:
            await self.update.reply(
                await self.format_message(
                    "passes-invitation-was-declined", couple=invitee
                ),
                reply_markup=InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                self.l("passes-button-solo"),
                                callback_data=f"{self.base.name}|solo|{self.pass_key}",
                            ),
                            InlineKeyboardButton(
                                self.l("passes-button-couple"),
                                callback_data=f"{self.base.name}|couple|{self.pass_key}",
                            ),
                        ],
                        [
                            InlineKeyboardButton(
                                self.l("passes-button-cancel"),
                                callback_data=f"{self.base.name}|pass_exit",
                            ),
                        ],
                    ]
                ),
                parse_mode=ParseMode.HTML,
            )

    def set_pass_key(self, pass_key: str) -> None:
        self.base.refresh_events_cache()
        if pass_key not in self.base.pass_keys:
            raise Exception(f"wrong pass key {pass_key}")
        self.pass_key = pass_key

    async def handle_cq_start(self, pass_key: str):
        self.set_pass_key(pass_key)
        user = await self.get_user()
        inviter_pass = await self.base.pass_db.find_one(
            {
                "bot_id": self.bot,
                "pass_key": self.pass_key,
                "couple": self.update.user,
            }
        )
        if inviter_pass is not None:
            inviter_user = await self.base.user_db.find_one(
                {
                    "user_id": inviter_pass.get("user_id"),
                    "bot_id": self.bot,
                }
            )
            inviter_id = inviter_pass.get("user_id")
            existing_pass = await self.get_pass()
            if (
                existing_pass is None
                or existing_pass.get("couple", 0) != inviter_id
            ):
                if existing_pass is None or existing_pass.get("state") != "paid":
                    return await self.show_couple_invitation(inviter_user or {"user_id": inviter_id})
                else:
                    inv_update = await self.base.create_update_from_user(inviter_id)
                    inv_update.set_pass_key(self.pass_key)
                    await inv_update.handle_invitation_declined(user)
        existing_pass = await self.get_pass()
        if existing_pass is not None:
            return await self.show_pass_edit(user, existing_pass)
        else:
            return await self.new_pass()

    async def handle_cq_pay(self, pass_key: str):
        pass_data = await self.get_pass(pass_key)
        if pass_data is None or pass_data.get("state") != "assigned":
            return await self.handle_cq_exit()
        await self.update.edit_or_reply(
            await self.format_message("passes-payment-request-callback-message"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )
        await self.update.reply(
            await self.format_message("passes-payment-request-waiting-message"),
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardMarkup(
                [[CANCEL_CHR + self.l("cancel-command")]], resize_keyboard=True
            ),
        )
        await self.update.require_anything(
            self.base.name,
            "handle_payment_proof_input",
            self.pass_key,
            "handle_payment_proof_timeout",
        )

    async def handle_cq_adm_acc(self, key: str, user_id_s: str):
        self.set_pass_key(key)
        if not self.base.is_payment_admin(self.update.user, self.pass_key):
            return
        user_id = int(user_id_s)
        pass_doc = await self.base.pass_db.find_one(
            {
                "bot_id": self.bot,
                "user_id": user_id,
                "pass_key": self.pass_key,
                "state": "paid",
            }
        )
        if pass_doc is None:
            return
        u_pass = self.base._pass_doc_to_data(pass_doc) or {}
        uids = [user_id]
        if "couple" in u_pass:
            uids.append(u_pass["couple"])
        accepted_ts = now_msk()
        await self.base.update_pass_fields(
            uids,
            self.pass_key,
            set_fields={
                "state": "paid",
                "proof_received": u_pass.get("proof_received"),
                "proof_file": u_pass.get("proof_file"),
                "proof_admin": u_pass.get("proof_admin"),
                "proof_admin_accepted": self.update.user,
                "proof_accepted": accepted_ts,
            },
        )
        await self.update.reply(
            await self.format_message(
                "passes-payment-proof-accepted",
                user=user_id,
                for_user=user_id,
            ),
            user_id,
            parse_mode=ParseMode.HTML,
        )
        await self.update.edit_message_text(
            await self.format_message(
                "passes-adm-payment-proof-accepted",
                user=user_id,
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_cq_adm_rej(self, key: str, user_id_s: str):
        self.set_pass_key(key)
        if not self.base.is_payment_admin(self.update.user, self.pass_key):
            return
        user_id = int(user_id_s)
        pass_doc = await self.base.pass_db.find_one(
            {
                "bot_id": self.bot,
                "user_id": user_id,
                "pass_key": self.pass_key,
                "state": "paid",
            }
        )
        if pass_doc is None:
            return
        u_pass = self.base._pass_doc_to_data(pass_doc) or {}
        uids = [user_id]
        if "couple" in u_pass:
            uids.append(u_pass["couple"])
        rejected_ts = now_msk()
        await self.base.update_pass_fields(
            uids,
            self.pass_key,
            set_fields={
                "state": "assigned",
                "proof_received": u_pass.get("proof_received"),
                "proof_file": u_pass.get("proof_file"),
                "proof_admin": u_pass.get("proof_admin"),
                "proof_rejected": rejected_ts,
            },
        )
        await self.update.reply(
            await self.format_message(
                "passes-payment-proof-rejected",
                user=user_id,
                for_user=user_id,
            ),
            user_id,
            parse_mode=ParseMode.HTML,
        )
        await self.update.edit_message_text(
            await self.format_message(
                "passes-adm-payment-proof-rejected",
                user=user_id,
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_payment_proof_timeout(self, data):
        return await self.update.reply(
            self.l("passes-payment-proof-timeout"),
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove(),
        )

    async def handle_legal_name_timeout(self, data):
        return await self.update.reply(
            self.l("passes-name-timeout"),
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove(),
        )

    async def handle_payment_proof_input(self, data):
        self.set_pass_key(data)
        if (
            filters.TEXT.check_update(self.tgUpdate)
            and self.update.message.text[0] == CANCEL_CHR
        ):
            return await self.update.reply(
                self.l("passes-payment-proof-cancelled"),
                parse_mode=ParseMode.HTML,
                reply_markup=ReplyKeyboardRemove(),
            )
        if not (filters.PHOTO | filters.Document.ALL).check_update(self.tgUpdate):
            return await self.update.reply(
                self.l(
                    "passes-payment-proof-wrong-data",
                ),
                parse_mode=ParseMode.HTML,
                reply_markup=ReplyKeyboardRemove(),
            )
        if filters.PHOTO.check_update(self.tgUpdate):
            doc = self.update.message.photo[-1]
            file_ext = ".jpg"
        else:
            doc = self.update.message.document
            import mimetypes

            file_ext = mimetypes.guess_extension(doc.mime_type)
        pass_data = await self.get_pass()
        if pass_data is None or pass_data.get("state") != "assigned":
            return await self.handle_cq_exit()
        uids = [self.update.user]
        req = {
            "bot_id": self.bot,
            "pass_key": self.pass_key,
            "state": "assigned",
            "type": pass_data.get("type", "solo"),
        }
        if "couple" in pass_data:
            uids.append(pass_data["couple"])
            req["couple"] = {"$in": uids}
        req["user_id"] = {"$in": uids}
        proof_received = now_msk()
        result = await self.base.pass_db.update_many(
            req,
            {
                "$set": {
                    "state": "paid",
                    "proof_received": proof_received,
                    "proof_file": f"{doc.file_id}{file_ext}",
                    "proof_admin": pass_data.get("proof_admin"),
                }
            },
        )
        if result.modified_count <= 0:
            return
        if pass_data.get("proof_admin") in self.base.get_all_payment_admins(
            self.pass_key
        ):
            admin = await self.base.user_db.find_one(
                {
                    "user_id": pass_data.get("proof_admin"),
                    "bot_id": self.bot,
                }
            )
            lc = "ru"
            if admin is not None and "language_code" in admin:
                lc = admin["language_code"]

            def l(s, **kwargs):  # noqa: E743
                return self.base.base_app.localization(s, args=kwargs, locale=lc)

            await self.update.forward_message(admin["user_id"])
            await self.update.reply(
                await self.format_message(
                    "passes-adm-payment-proof-received",
                    user=self.update.user,
                    for_user=admin["user_id"],
                ),
                chat_id=admin["user_id"],
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                l("passes-adm-payment-proof-accept-button"),
                                callback_data=f"{self.base.name}|adm_acc|{self.pass_key}|{self.update.user}",
                            ),
                            InlineKeyboardButton(
                                l("passes-adm-payment-proof-reject-button"),
                                callback_data=f"{self.base.name}|adm_rej|{self.pass_key}|{self.update.user}",
                            ),
                        ]
                    ]
                ),
            )
        await self.update.reply(
            self.l("passes-payment-proof-forwarded"),
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove(),
        )

    async def new_pass(self):
        event = self.base.require_event(self.pass_key)
        if now_msk() < event.sell_start:
            return await self.update.edit_or_reply(
                self.l_pass("passes-sell-not-started"),
                reply_markup=InlineKeyboardMarkup([]),
                parse_mode=ParseMode.HTML,
            )
        await self.update.edit_or_reply(
            self.l_pass("passes-pass-create-start-message"),
            reply_markup=InlineKeyboardMarkup([]),
            parse_mode=ParseMode.HTML,
        )
        if not self.is_passport_required():
            return await self.after_legal_name_input("pass", "skip")
        return await self.handle_name_request("pass")

    async def handle_start(self):
        logger.debug(f"starting passes for: {self.update.user}")
        self.base.refresh_events_cache()
        buttons = []
        for pass_key in self.base.pass_keys:
            buttons.append(
                InlineKeyboardButton(
                    self.l_pass("passes-select-type-button", pass_key=pass_key),
                    callback_data=f"{self.base.name}|start|{pass_key}",
                )
            )
        buttons.append(
            InlineKeyboardButton(
                self.l("passes-button-exit"), callback_data=f"{self.base.name}|exit"
            )
        )
        await self.update.reply(
            self.l("passes-select-type-message"),
            reply_markup=InlineKeyboardMarkup([buttons]),
            parse_mode=ParseMode.HTML,
        )

    async def handle_cq_name(self):
        logger.debug(f"command to change legal name for: {self.update.user}")
        return await self.handle_name_request("cmd")

    async def handle_cq_couple(self, key: str):
        self.set_pass_key(key)
        existing_pass = await self.get_pass()
        if isinstance(existing_pass, dict):
            state = existing_pass.get("state")
            is_waitlist_solo = (
                state == "waitlist"
                and existing_pass.get("type", "solo") == "solo"
                and "couple" not in existing_pass
            )
            if not is_waitlist_solo and state != "waiting-for-couple":
                return await self.handle_cq_exit()
        markup = ReplyKeyboardMarkup(
            [[CANCEL_CHR + self.l("cancel-command")]], resize_keyboard=True
        )
        try:
            await self.update.edit_message_text(
                self.l_pass("passes-couple-request-edit"),
                reply_markup=InlineKeyboardMarkup([]),
                parse_mode=ParseMode.HTML,
            )
        except Exception as err:
            logger.error(
                f"Exception in handle_cq_couple: {err=}, {self.update.user=}",
                exc_info=1,
            )
        await self.update.reply(
            self.l_pass("passes-couple-request-message"),
            reply_markup=markup,
            parse_mode=ParseMode.HTML,
        )
        await self.update.require_anything(
            self.base.name,
            "handle_couple_input",
            self.pass_key,
            "handle_couple_timeout",
        )

    async def handle_couple_timeout(self, _data):
        await self.update.reply(
            self.l(
                "passes-couple-request-timeout",
            ),
            reply_markup=InlineKeyboardMarkup([]),
            parse_mode=ParseMode.HTML,
        )

    async def handle_cq_solo(self, key: str):
        self.set_pass_key(key)
        user = await self.get_user()
        existing_pass = await self.base.get_pass_for_user(self.update.user, self.pass_key)
        if existing_pass is None and self.pass_key in user and isinstance(user[self.pass_key], dict):
            existing_pass = user[self.pass_key]
        if isinstance(existing_pass, dict):
            state = existing_pass.get("state")
            is_waitlist_solo = (
                state == "waitlist"
                and existing_pass.get("type", "solo") == "solo"
                and "couple" not in existing_pass
            )
            if not is_waitlist_solo and state != "waiting-for-couple":
                return await self.handle_cq_exit()
        u_pass = {
            "role": user["role"],
            "state": "waitlist",
            "date_created": now_msk(),
            "type": "solo",
            "proof_admin": user.get("proof_admins", {}).get(
                self.pass_key,
                self.base.pick_payment_admin(self.pass_key),
            ),
        }
        if isinstance(existing_pass, dict):
            u_pass["date_created"] = existing_pass.get("date_created", u_pass["date_created"])
            u_pass["role"] = existing_pass.get("role", u_pass["role"])
            u_pass["proof_admin"] = existing_pass.get(
                "proof_admin",
                u_pass["proof_admin"],
            )
            if u_pass["proof_admin"] not in self.base.payment_admins[self.pass_key]:
                u_pass["proof_admin"] = self.base.pick_payment_admin(self.pass_key)
        await self.base.save_pass_data(self.update.user, self.pass_key, u_pass)
        keys = dict(u_pass)
        if self.is_passport_required():
            keys["name"] = user.get("legal_name", "")
        await self.update.edit_or_reply(
            self.l_pass("passes-solo-saved"),
            reply_markup=InlineKeyboardMarkup([]),
            parse_mode=ParseMode.HTML,
        )

        await self.base.recalculate_queues()

    async def handle_couple_input(self, data):
        self.set_pass_key(data)
        if (
            filters.TEXT.check_update(self.tgUpdate)
            and self.update.message.text[0] == CANCEL_CHR
        ):
            return await self.update.reply(
                self.l("passes-couple-request-cancelled"),
                parse_mode=ParseMode.HTML,
                reply_markup=ReplyKeyboardRemove(),
            )
        msg = self.update.update.message
        if (
            msg is None
            or (
                msg.forward_origin is None
                and (
                    msg.contact is None
                    or not isinstance(msg.contact, Contact)
                    or msg.contact.user_id is None
                    or msg.contact.user_id == self.update.user
                )
            )
            or not isinstance(msg.forward_origin, MessageOriginUser)
            or msg.forward_origin.sender_user.id == self.update.user
        ):
            return await self.update.reply(
                self.l("passes-couple-request-wrong-data"),
                reply_markup=ReplyKeyboardRemove(),
                parse_mode=ParseMode.HTML,
            )
        other_user_id = msg.forward_origin.sender_user.id

        invitee_pass = await self.base.get_pass_for_user(other_user_id, self.pass_key)
        invitee = None
        if invitee_pass is None:
            invitee = await self.base.user_db.find_one(
                {
                    "user_id": other_user_id,
                    "bot_id": self.update.bot.id,
                }
            )
            if invitee is not None and self.pass_key in invitee:
                invitee_pass = invitee[self.pass_key]
        if isinstance(invitee_pass, dict) and invitee_pass.get("state") == "paid":
            return await self.update.reply(
                self.l_pass("passes-couple-request-invitee-paid"),
                reply_markup=ReplyKeyboardRemove(),
                parse_mode=ParseMode.HTML,
            )

        user = await self.get_user()
        existing_pass = await self.base.get_pass_for_user(self.update.user, self.pass_key)
        if existing_pass is None and self.pass_key in user and isinstance(user[self.pass_key], dict):
            existing_pass = user[self.pass_key]
        u_pass = {
            "role": user["role"],
            "state": "waiting-for-couple",
            "date_created": now_msk(),
            "type": "couple",
            "couple": other_user_id,
            "proof_admin": user.get("proof_admins", {}).get(
                self.pass_key,
                self.base.pick_payment_admin(self.pass_key),
            ),
        }
        if isinstance(existing_pass, dict):
            u_pass["date_created"] = existing_pass.get("date_created", u_pass["date_created"])
            u_pass["role"] = existing_pass.get("role", u_pass["role"])
            u_pass["proof_admin"] = existing_pass.get(
                "proof_admin",
                u_pass["proof_admin"],
            )
            if u_pass["proof_admin"] not in self.base.payment_admins[self.pass_key]:
                u_pass["proof_admin"] = self.base.pick_payment_admin(self.pass_key)
        await self.base.save_pass_data(self.update.user, self.pass_key, u_pass)
        keys = dict(u_pass)
        if self.is_passport_required():
            keys["name"] = user.get("legal_name", "")
        if invitee is not None:
            inv_update = await self.base.create_update_from_user(invitee["user_id"])
            inv_update.set_pass_key(self.pass_key)
            await inv_update.show_couple_invitation(user)
            await self.update.reply(
                self.l_pass("passes-couple-saved-sent"),
                reply_markup=ReplyKeyboardRemove(),
                parse_mode=ParseMode.HTML,
            )
        else:
            await self.update.reply(
                self.l_pass("passes-couple-saved"),
                reply_markup=ReplyKeyboardRemove(),
                parse_mode=ParseMode.HTML,
            )

    async def handle_name_request(self, data):
        user = await self.get_user()
        if "legal_name_frozen" in user:
            return await self.after_legal_name_input(data)
        btns = []
        if "legal_name" in user:
            btns.append([user["legal_name"]])
        btns.append([CANCEL_CHR + self.l("cancel-command")])
        markup = ReplyKeyboardMarkup(btns, resize_keyboard=True)
        await self.update.reply(
            self.l("passes-legal-name-request-message"),
            reply_markup=markup,
            parse_mode=ParseMode.HTML,
        )
        await self.update.require_input(
            self.base.name,
            "handle_legal_name_input",
            f"{self.pass_key}|{data}",
            "handle_legal_name_timeout",
        )
    
    async def require_passport_data(self):
        await self.update.reply(
            self.l("passes-passport-data-required"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(
                    self.l("passes-passport-data-button"),
                    callback_data=f"{self.base.name}|passport_data|{self.pass_key}"
                )
            ]]),
        )

    async def handle_passport_data_cmd(self):
        return await self.handle_cq_passport_data(self.base.default_pass_key())

    async def handle_cq_passport_data(self, pass_key: str):
        self.set_pass_key(pass_key)
        logger.debug(f"command to change passport data for: {self.update.user}")
        if self.pass_key not in self.base.pass_keys:
            return await self.update.reply(
                self.l("passes-passport-data-required-beginning-message"),
                parse_mode=ParseMode.HTML,
                reply_markup=ReplyKeyboardRemove(),
            )
        return await self.handle_name_request("passport_data")

    async def handle_passport_request(self, data: str):
        user = await self.get_user()
        if "legal_name_frozen" in user:
            return await self.after_legal_name_input(data)
        btns = []
        if "passport_number" in user:
            btns.append([user["passport_number"]])
        btns.append([CANCEL_CHR + self.l("cancel-command")])
        markup = ReplyKeyboardMarkup(btns, resize_keyboard=True)
        await self.update.reply(
            self.l("passes-passport-request-message"),
            reply_markup=markup,
            parse_mode=ParseMode.HTML,
        )
        await self.update.require_input(
            self.base.name,
            "handle_passport_input",
            f"{self.pass_key}|{data}",
            "handle_passport_timeout",
        )
    
    async def handle_passport_input(self, data: str):
        pass_key, reason = data.split("|", maxsplit=2)
        if pass_key != "":
            self.set_pass_key(pass_key)
        user = await self.get_user()
        passport_number = self.update.message.text
        if passport_number[0] == CANCEL_CHR:
            return await self.update.reply(
                self.l("passes-pass-create-cancel"),
                parse_mode=ParseMode.HTML,
                reply_markup=ReplyKeyboardRemove(),
            )
        old_passport = user.get("passport_number", "")
        logger.debug(
            f"changing passport number for: {self.update.user} from {old_passport} to {passport_number}"
        )
        await self.base.user_db.update_one(
            {"user_id": self.update.user, "bot_id": self.bot},
            {"$set": {"passport_number": passport_number}},
        )
        await self.update.reply(
            self.l(
                "passes-passport-changed-message",
                passportNumber=passport_number,
            ),
            reply_markup=ReplyKeyboardRemove(),
            parse_mode=ParseMode.HTML,
        )
        await self.after_legal_name_input(reason, "passport")
    
    async def handle_passport_timeout(self, data):
        return await self.update.reply(
            self.l("passes-passport-timeout"),
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove(),
        )

    async def handle_legal_name_input(self, data: str):
        pass_key, reason = data.split("|", maxsplit=2)
        if pass_key != "":
            self.set_pass_key(pass_key)
        user = await self.get_user()
        name = self.update.message.text
        if name[0] == CANCEL_CHR:
            return await self.update.reply(
                self.l("passes-pass-create-cancel"),
                parse_mode=ParseMode.HTML,
                reply_markup=ReplyKeyboardRemove(),
            )
        old_name = user["legal_name"] if "legal_name" in user else ""
        logger.debug(
            f"changing legal name for: {self.update.user} from {old_name} to {name}"
        )
        await self.base.user_db.update_one(
            {"user_id": self.update.user, "bot_id": self.bot},
            {"$set": {"legal_name": name}},
        )
        await self.update.reply(
            self.l(
                "passes-legal-name-changed-message",
                name=name,
            ),
            reply_markup=ReplyKeyboardRemove(),
            parse_mode=ParseMode.HTML,
        )
        await self.after_legal_name_input(reason)

    async def after_legal_name_input(self, data, input_type: str = "name"):
        if data == "cmd":
            return
        if input_type == "name" and self.is_passport_required():
            return await self.handle_passport_request(data)
        if data == "pass":
            await self.update.reply(
                self.l("passes-pass-role-select"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                self.l("passes-role-button-leader"),
                                callback_data=f"{self.base.name}|pass_role|{self.pass_key}|l",
                            ),
                            InlineKeyboardButton(
                                self.l("passes-role-button-follower"),
                                callback_data=f"{self.base.name}|pass_role|{self.pass_key}|f",
                            ),
                        ],
                        [
                            InlineKeyboardButton(
                                self.l("passes-role-button-cancel"),
                                callback_data=f"{self.base.name}|pass_exit",
                            ),
                        ],
                    ]
                ),
            )
        elif data not in ["cmd", "passport_data"]:
            await self.accept_invitation(data)

    async def handle_cq_pass_role(self, key: str, role: str):
        self.set_pass_key(key)
        new_role = "leader" if role.startswith("l") else "follower"
        setter = {
            "role": new_role,
        }
        if len(self.base.payment_admins[self.pass_key]) == 1:
            setter["proof_admins." + self.pass_key] = \
                self.base.payment_admins[self.pass_key][0]
        await self.base.user_db.update_one(
            {"user_id": self.update.user, "bot_id": self.bot},
            {"$set": setter},
        )

        if len(self.base.payment_admins[self.pass_key]) > 1:
            await self.handle_cq_change_admin(self.pass_key, "passes-pass-role-saved")
        else:
            await self.update.edit_or_reply(
                self.l("passes-pass-admin-saved"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                self.l("passes-button-solo"),
                                callback_data=f"{self.base.name}|solo|{self.pass_key}",
                            ),
                            InlineKeyboardButton(
                                self.l("passes-button-couple"),
                                callback_data=f"{self.base.name}|couple|{self.pass_key}",
                            ),
                        ],
                        [
                            InlineKeyboardButton(
                                self.l("passes-button-cancel"),
                                callback_data=f"{self.base.name}|pass_exit",
                            ),
                        ],
                    ]
                ),
            )

    async def handle_cq_role(self, role: str):
        new_role = "leader" if role.startswith("l") else "follower"
        await self.base.user_db.update_one(
            {"user_id": self.update.user, "bot_id": self.bot},
            {"$set": {"role": new_role}},
        )

        await self.update.edit_or_reply(
            self.l(
                "passes-role-saved",
                role=new_role,
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def handle_cq_role_exit(self):
        await self.update.edit_or_reply(
            self.l("passes-role-exit"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def get_user(self):
        if self.user is None:
            self.user = await self.update.get_user()
        if self.pass_key and self.pass_data is None and self.pass_owner_id is None:
            cached = await self.base.get_pass_for_user(self.update.user, self.pass_key)
            if cached is not None:
                self.pass_data = cached
                self.pass_owner_id = self.update.user
        return self.user

    async def get_pass(self, pass_key: str | None = None, user_id: int | None = None):
        if pass_key is not None:
            self.set_pass_key(pass_key)
        else:
            pass_key = self.pass_key
        assert pass_key is not None, "pass_key must be set"
        uid = user_id or self.update.user
        if self.pass_data is not None and self.pass_owner_id == uid:
            return self.pass_data
        pass_doc = await self.base.get_pass_for_user(uid, pass_key)
        if pass_doc is not None and uid == self.update.user:
            self.pass_data = pass_doc
            self.pass_owner_id = uid
        return pass_doc

    async def assign_pass(
        self,
        pass_key: str,
        price: int | None = None,
        type: str | None = None,
        comment: str | None = None,
        skip_in_balance_count: bool = False,
        proof_admin: int | None = None,
        price_by_user: dict[int, int] | None = None,
        pass_type_index_by_user: dict[int, int] | None = None,
    ):
        self.set_pass_key(pass_key)
        pass_doc = await self.base.pass_db.find_one(
            {
                "bot_id": self.bot,
                "user_id": self.update.user,
                "pass_key": self.pass_key,
                "state": "waitlist",
            }
        )
        if pass_doc is None:
            return False

        pass_data = self.base._pass_doc_to_data(pass_doc) or {}
        uids = [self.update.user]
        if "couple" in pass_data:
            uids.append(pass_data["couple"])
        is_couple = "couple" in pass_data and len(uids) > 1
        resolved_pass_type_index_by_user: dict[int, int] | None = (
            dict(pass_type_index_by_user) if pass_type_index_by_user is not None else None
        )

        if price_by_user is not None:
            missing_uids = [uid for uid in uids if uid not in price_by_user]
            if missing_uids:
                raise ValueError(
                    f"price_by_user is missing values for {missing_uids}"
                )
            prices = {uid: int(price_by_user[uid]) for uid in uids}
        elif price is None:
            tier_pricing = await self.base.resolve_candidate_tier_prices(
                self.pass_key,
                self.update.user,
                pass_data,
            )
            if tier_pricing is None:
                logger.warning(
                    "Cannot resolve current tier pricing for admin assignment: "
                    "user=%s pass=%s",
                    self.update.user,
                    self.pass_key,
                )
                return False
            prices, resolved_pass_type_index_by_user = tier_pricing
        else:
            total_price = self.base.resolve_total_pass_price(
                self.pass_key,
                uids,
                custom_total_price=price,
            )
            prices = self.base.split_total_price(total_price, uids)

        state_by_user = {
            uid: ("paid" if prices.get(uid, 0) == 0 else "assigned")
            for uid in uids
        }
        split_couple_for_free = (
            is_couple
            and any(prices.get(uid, 0) == 0 for uid in uids)
        )
        assignment_ts = now_msk()
        update_matched_count = 0
        have_changes = False
        for uid in uids:
            set_fields: dict[str, object] = {
                "state": state_by_user.get(uid, "assigned"),
                "date_assignment": assignment_ts,
            }
            if comment is not None:
                set_fields["comment"] = comment
            if split_couple_for_free:
                set_fields["type"] = "solo"
            elif type is not None:
                set_fields["type"] = type
            if skip_in_balance_count:
                set_fields["skip_in_balance_count"] = True

            if state_by_user.get(uid) == "paid":
                free_proof_admin = proof_admin
                if free_proof_admin is None:
                    free_proof_admin = pass_data.get("proof_admin")
                if free_proof_admin is None:
                    free_proof_admin = self.base.pick_payment_admin(self.pass_key)
                set_fields["proof_received"] = assignment_ts
                set_fields["proof_file"] = "free_pass"
                set_fields["proof_admin"] = free_proof_admin
                set_fields["proof_accepted"] = assignment_ts

            match_fields: dict[str, object] = {
                "bot_id": self.bot,
                "pass_key": self.pass_key,
                "user_id": uid,
                "state": "waitlist",
            }
            if is_couple:
                match_fields["couple"] = {"$in": uids}
            set_fields["price"] = int(prices.get(uid, 0))
            if (
                resolved_pass_type_index_by_user is not None
                and uid in resolved_pass_type_index_by_user
            ):
                tier_index = int(resolved_pass_type_index_by_user[uid])
                set_fields["pass_type_index"] = {"$ifNull": ["$pass_type_index", tier_index]}
                set_fields["assignment_tier_number"] = {
                    "$ifNull": ["$assignment_tier_number", tier_index + 1]
                }
            update_pipeline: list[dict[str, object]] = [{"$set": set_fields}]
            if split_couple_for_free:
                update_pipeline.append({"$unset": "couple"})

            result = await self.base.pass_db.update_one(match_fields, update_pipeline)
            update_matched_count += result.matched_count
            if result.modified_count > 0:
                have_changes = True

        current_count = await self.base.pass_db.count_documents(
            {"bot_id": self.bot, "pass_key": self.pass_key, "user_id": {"$in": uids}}
        )
        if update_matched_count < len(uids) or current_count < len(uids):
            await self.base.delete_passes(uids, self.pass_key)
            have_changes = True

        users = []
        async for doc in self.base.user_db.find({"bot_id": self.bot, "user_id": {"$in": uids}}):
            users.append(doc)
        pass_docs_after = {
            d["user_id"]: self.base._pass_doc_to_data(d)
            for d in await self.base.pass_db.find(
                {"bot_id": self.bot, "pass_key": self.pass_key, "user_id": {"$in": uids}}
            ).to_list(None)
        }
        for user_doc in users:
            uid = user_doc["user_id"]
            if current_count < len(uids):
                if len(uids) > 1:
                    upd = await self.base.create_update_from_user(uid)
                    await upd.update.reply(
                        upd.l("passes-error-couple-not-found"),
                        parse_mode=ParseMode.HTML,
                    )
            else:
                pass_info = pass_docs_after.get(uid, pass_data)
                logger.info(
                    f"assigned pass to {uid}, role {pass_info.get('role')}, name {user_doc.get('legal_name', client_user_name(user_doc))}"
                )
                upd = await self.base.create_update_from_user(uid)
                upd.set_pass_key(self.pass_key)
                await upd.show_pass_edit(
                    user_doc,
                    pass_info,
                    (
                        "passes-pass-free-assigned"
                        if prices.get(uid, 0) == 0
                        else "passes-pass-assigned"
                    ),
                )
        return have_changes

    async def handle_role_cmd(self):
        logger.debug(f"setting role for: {self.update.user}")
        user = await self.get_user()
        text = self.l("passes-role-select")
        if "role" in user:
            text = self.l(
                "passes-role-change-select",
                role=user["role"],
            )

        await self.update.edit_or_reply(
            text,
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(
                            self.l("passes-role-button-leader"),
                            callback_data=f"{self.base.name}|role|l",
                        ),
                        InlineKeyboardButton(
                            self.l("passes-role-button-follower"),
                            callback_data=f"{self.base.name}|role|f",
                        ),
                    ],
                    [
                        InlineKeyboardButton(
                            self.l("passes-role-button-cancel"),
                            callback_data=f"{self.base.name}|role_exit",
                        ),
                    ],
                ]
            ),
        )

    async def inform_pass_cancelled(self):
        await self.update.reply(
            await self.format_message("passes-pass-cancelled-by-other"),
            parse_mode=ParseMode.HTML,
        )

    async def cancel_pass(self) -> bool:
        pass_data = await self.get_pass()
        if not isinstance(pass_data, dict):
            return False
        if pass_data.get("state") == "paid":
            return False
        uids = [self.update.user]
        if (
            "couple" in pass_data
            and pass_data.get("state") != "waiting-for-couple"
        ):
            uids.append(pass_data["couple"])
            upd = await self.base.create_update_from_user(pass_data["couple"])
            upd.set_pass_key(self.pass_key)
            await upd.inform_pass_cancelled()
        delete_result = await self.base.pass_db.delete_many(
            {
                "bot_id": self.bot,
                "pass_key": self.pass_key,
                "user_id": {"$in": uids},
                "state": {"$ne": "paid"},
            }
        )
        return delete_result.deleted_count > 0

    async def notify_no_more_passes(self, pass_key):
        self.set_pass_key(pass_key)
        ts = now_msk()
        pass_data = await self.get_pass()
        if pass_data is None or pass_data.get("state") != "waitlist":
            return
        await self.base.update_pass_fields(
            [self.update.user],
            self.pass_key,
            set_fields={"no_more_passes_notification_sent": ts},
        )
        await self.show_pass_edit(
            await self.get_user(),
            {**pass_data, "no_more_passes_notification_sent": ts},
            "passes-added-to-waitlist",
        )

    async def notify_deadline_close(self, suffix: str = ""):
        ts = now_msk()
        pass_data = await self.get_pass()
        if pass_data is None or pass_data.get("state") != "assigned":
            return
        if ("notified_deadline_close" + suffix) in pass_data:
            return
        await self.base.update_pass_fields(
            [self.update.user],
            self.pass_key,
            set_fields={"notified_deadline_close" + suffix: ts},
        )
        await self.update.reply(
            self.l_pass("passes-payment-deadline-close"),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([]),
        )

    async def decline_due_deadline(self, pass_key):
        self.set_pass_key(pass_key)
        pass_data = await self.get_pass()
        if pass_data is None or "couple" not in pass_data:
            return
        await self.base.update_pass_fields(
            [self.update.user],
            self.pass_key,
            unset_fields=["couple"],
        )
        await self.update.reply(
            await self.format_message("passes-couple-didnt-answer"),
            reply_markup=InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(
                            self.l("passes-button-solo"),
                            callback_data=f"{self.base.name}|solo|{self.pass_key}",
                        ),
                        InlineKeyboardButton(
                            self.l("passes-button-couple"),
                            callback_data=f"{self.base.name}|couple|{self.pass_key}",
                        ),
                    ],
                    [
                        InlineKeyboardButton(
                            self.l("passes-button-cancel"),
                            callback_data=f"{self.base.name}|pass_exit",
                        ),
                    ],
                ]
            ),
            parse_mode=ParseMode.HTML,
        )

    async def decline_invitation_due_deadline(self, pass_key):
        self.set_pass_key(pass_key)
        await self.update.reply(
            await self.format_message("passes-invitation-timeout"),
            parse_mode=ParseMode.HTML,
        )

    async def cancel_due_deadline(self, pass_key):
        self.set_pass_key(pass_key)
        success = await self.cancel_pass()
        if success:
            await self.update.reply(
                self.l("passes-payment-deadline-exceeded"),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup([]),
            )

    async def handle_passes_assign(self):
        assert self.update.user in self.base.config.telegram.admins, (
            f"{self.update.user} is not admin"
        )
        self.base.refresh_events_cache()
        args_list = self.update.parse_cmd_arguments()

        parser = SilentArgumentParser()
        parser.add_argument("--pass_key", type=str, help="Pass key")
        parser.add_argument("--price", type=int, help="Custom price")
        parser.add_argument("--type", type=str, help="Custom type")
        parser.add_argument("--create_last", action="store_true", help="Create using last settings")
        parser.add_argument(
            "--leader",
            action="store_true",
            help="Create if not exists, with role leader",
        )
        parser.add_argument(
            "--follower",
            action="store_true",
            help="Create if not exists, with role follower",
        )
        parser.add_argument(
            "--create_name", type=str, help="Create with this legal name"
        )
        parser.add_argument("--skip", action="store_true", help="Skip in balance count")
        parser.add_argument("--comment", type=str, help="Comment message")
        parser.add_argument("recipients", nargs="*", help="Recipients")

        args = parser.parse_args(args_list[1:])
        logger.debug(f"passes_assign {args=}, {args_list=}")
        assert args.pass_key in self.base.pass_keys, f"wrong pass key {args.pass_key}"
        assert args.price is None or args.price >= 0, f"wrong price {args.price}"
        assigned = []
        for recipient in args.recipients:
            try:
                user_id = int(recipient)
                created_pass_data = None
                pass_doc = await self.base.pass_db.find_one(
                    {
                        "user_id": user_id,
                        "bot_id": self.bot,
                        "pass_key": args.pass_key,
                        "state": "waitlist",
                    }
                )
                if pass_doc is None:
                    if ((
                        args.leader is not None or args.follower is not None
                    ) and args.create_name is not None):
                        created_pass_data = {
                            "state": "waitlist",
                            "type": "solo",
                            "proof_admin": self.base.payment_admins[args.pass_key][0],
                            "role": "leader" if args.leader else "follower",
                            "date_created": now_msk(),
                        }
                    elif args.create_last:
                        last_role = None
                        user = await self.base.user_db.find_one(
                            {
                                "user_id": user_id,
                                "bot_id": self.bot,
                            }
                        )
                        if user is not None:
                            last_role = user.get("role")
                            if last_role is None:
                                for event_key in self.base.all_pass_keys():
                                    embedded_pass = user.get(event_key)
                                    if isinstance(embedded_pass, dict):
                                        last_role = embedded_pass.get("role")
                                        if last_role is not None:
                                            break
                        if last_role is None:
                            last_pass = await self.base.pass_db.find_one(
                                {"bot_id": self.bot, "user_id": user_id},
                                sort=[("date_created", -1)],
                            )
                            if last_pass is not None:
                                last_pass_data = self.base._pass_doc_to_data(last_pass) or {}
                                last_role = last_pass_data.get("role")
                        if last_role is None:
                            raise ValueError(
                                f"user {user_id} has no last role to create pass from"
                            )
                        created_pass_data = {
                            "state": "waitlist",
                            "type": "solo",
                            "proof_admin": self.base.pick_payment_admin(args.pass_key),
                            "role": last_role,
                            "date_created": now_msk(),
                        }
                    else:
                        raise ValueError(
                            f"user {user_id} with pass {args.pass_key} not found"
                        )
                if created_pass_data is not None:
                    if args.create_name is not None:
                        await self.base.user_db.update_one(
                            {"user_id": user_id, "bot_id": self.bot},
                            {"$set": {"legal_name": args.create_name}},
                        )
                    await self.base.save_pass_data(
                        user_id, args.pass_key, created_pass_data
                    )
                upd = await self.base.create_update_from_user(user_id)

                assigned_ok = await upd.assign_pass(
                    args.pass_key, args.price, args.type, args.comment, args.skip, self.update.user
                )
                if assigned_ok:
                    logger.info(f"pass {args.pass_key=} assigned to {recipient=}")
                    assigned.append(user_id)
                else:
                    logger.warning(
                        "pass %s was not assigned to recipient=%s",
                        args.pass_key,
                        recipient,
                    )
            except Exception as err:
                logger.error(f"passes_assign {err=}, {recipient=}", exc_info=1)
        await self.update.reply(
            f"passes_assign done: {assigned}", parse_mode=ParseMode.HTML
        )
        await self.base.recalculate_queues()

    async def handle_passes_tier(self):
        args_list = self.update.parse_cmd_arguments()
        self.base.refresh_events_cache()

        parser = SilentArgumentParser()
        parser.add_argument(
            "--pass_key",
            type=str,
            help="Pass key",
            default=self.base.default_pass_key(),
        )
        args = parser.parse_args(args_list[1:])
        logger.debug(f"passes_tier {args=}, {args_list=}")
        assert args.pass_key in self.base.pass_keys, f"wrong pass key {args.pass_key}"
        self.set_pass_key(args.pass_key)
        assert (
            self.update.user in self.base.config.telegram.admins
            or self.base.is_payment_admin(self.update.user, self.pass_key)
        ), f"{self.update.user} is not admin"
        await self.update.reply(
            await self.base.format_current_tier_status(self.pass_key),
            parse_mode=None,
        )
    
    async def handle_passes_uncouple(self):
        """Admin command: /passes_uncouple <pass_key> <user_id>
        Splits a couple pass into two solo passes for the specified user and their partner.
        Logic:
          - Validate admin rights
          - Load user doc with couple pass (type == couple)
          - Load coupled user doc; ensure both have pass entries
          - Keep each stored per-user price as-is
          - Update both passes: set type solo, remove couple link
        """
        assert self.update.user in self.base.config.telegram.admins, (
            f"{self.update.user} is not admin"
        )
        args_list = self.update.parse_cmd_arguments()
        self.base.refresh_events_cache()
        # Expect: ['/passes_uncouple', pass_key, user_id]
        if len(args_list) < 3:
            return await self.update.reply(
                "Usage: /passes_uncouple <pass_key> <user_id>",
                parse_mode=None,
            )
        pass_key = args_list[1]
        if pass_key not in self.base.pass_keys:
            return await self.update.reply(
                f"Unknown pass_key {pass_key}", parse_mode=None
            )
        self.set_pass_key(pass_key)
        try:
            target_user_id = int(args_list[2])
        except ValueError:
            return await self.update.reply(
                f"Invalid user id: {args_list[2]}", parse_mode=None
            )

        # Load target user with couple pass
        user, user_pass = await self.base.get_user_with_pass(target_user_id, pass_key)
        if user is None or not isinstance(user_pass, dict) or user_pass.get("type") != "couple":
            return await self.update.reply(
                f"User {target_user_id} does not have a couple pass {pass_key}",
                parse_mode=None,
            )
        couple_id = user_pass.get("couple")
        if couple_id is None:
            return await self.update.reply(
                f"User {target_user_id} does not have a couple partner recorded", parse_mode=None
            )
        couple_user, couple_pass = await self.base.get_user_with_pass(couple_id, pass_key)
        if couple_user is None or not isinstance(couple_pass, dict) or couple_pass.get("type") != "couple":
            return await self.update.reply(
                f"Couple user {couple_id} does not have a matching couple pass", parse_mode=None
            )

        # Update both users
        updated_first = dict(user_pass)
        updated_first["type"] = "solo"
        updated_first.pop("couple", None)
        updated_second = dict(couple_pass)
        updated_second["type"] = "solo"
        updated_second.pop("couple", None)
        await self.base.save_pass_data(target_user_id, pass_key, updated_first)
        await self.base.save_pass_data(couple_id, pass_key, updated_second)
        await self.update.reply(
            f"Uncoupled users {target_user_id} and {couple_id} for pass {pass_key}. Prices were preserved per user.",
            parse_mode=None,
        )
        await self.base.recalculate_queues()

    async def handle_passes_cancel(self):
        args_list = self.update.parse_cmd_arguments()
        self.base.refresh_events_cache()

        parser = SilentArgumentParser()
        parser.add_argument(
            "--pass_key",
            type=str,
            help="Pass key",
            default=self.base.default_pass_key(),
        )
        parser.add_argument("recipients", nargs="*", help="Recipients")

        args = parser.parse_args(args_list[1:])
        logger.debug(f"passes_cancel {args=}, {args_list=}")
        assert args.pass_key in self.base.pass_keys, f"wrong pass key {args.pass_key}"
        self.set_pass_key(args.pass_key)
        assert (self.update.user in self.base.config.telegram.admins or
                self.base.is_payment_admin(self.update.user, self.pass_key)), (
            f"{self.update.user} is not admin"
        )
        cancelled = []
        recipients = [int(recipient) for recipient in args.recipients]
        recipients_set = set(recipients)
        for recipient in recipients:
            try:
                pass_doc = await self.base.pass_db.find_one(
                    {
                        "bot_id": self.bot,
                        "user_id": recipient,
                        "pass_key": self.pass_key,
                    }
                )
                if pass_doc is None:
                    continue
                pass_data = self.base._pass_doc_to_data(pass_doc) or {}
                if pass_data.get("type") == "couple":
                    partner_id = pass_data.get("couple")
                    if partner_id is not None and int(partner_id) not in recipients_set:
                        result = await self.base.pass_db.update_one(
                            {
                                "bot_id": self.bot,
                                "user_id": partner_id,
                                "pass_key": self.pass_key,
                                "couple": {"$eq": recipient},
                            },
                            {
                                "$unset": {
                                    "couple": "",
                                },
                                "$set": {
                                    "type": "solo",
                                },
                            },
                        )
                        if result.modified_count > 0:
                            logger.info(
                                f"pass {args.pass_key=} for {recipient=} couple {partner_id} was changed to solo"
                            )
                        else:
                            logger.error(
                                f"pass {args.pass_key=} for {recipient=} couple {partner_id} was not changed to solo"
                            )
                result = await self.base.pass_db.delete_one(
                    {
                        "bot_id": self.bot,
                        "user_id": recipient,
                        "pass_key": self.pass_key,
                    }
                )
                if result.deleted_count <= 0:
                    logger.info(
                        f"pass {self.pass_key=} for {recipient=} was not cancelled"
                    )
                else:
                    logger.info(f"pass {self.pass_key=} cancelled for {recipient=}")
                    cancelled.append(recipient)
            except Exception as err:
                logger.error(f"passes_cancel {err=}, {recipient=}", exc_info=1)
        await self.update.reply(
            f"passes_cancel done: {cancelled}", parse_mode=ParseMode.HTML
        )
        await self.base.recalculate_queues()
    
    async def handle_passes_switch_to_me_cmd(self):
        args_list = self.update.parse_cmd_arguments()
        self.base.refresh_events_cache()
        tail = args_list[1:]
        if len(tail) == 0:
            return await self.update.reply(
                "Usage: /passes_switch_to_me [<pass_key>=active] <user_id>",
                parse_mode=None,
            )
        pass_key = self.base.default_pass_key()
        target_user_raw = None
        if len(tail) == 1:
            target_user_raw = tail[0]
        else:
            if tail[0] in self.base.pass_keys:
                pass_key = tail[0]
                target_user_raw = tail[1]
            else:
                target_user_raw = tail[0]
        if pass_key not in self.base.pass_keys:
            return await self.update.reply(
                f"Unknown pass key {pass_key}", parse_mode=None
            )
        try:
            target_user_id = int(target_user_raw)
        except Exception:
            return await self.update.reply(
                "Invalid user id", parse_mode=None,
            )
        self.set_pass_key(pass_key)
        if not (
            self.update.user in self.base.config.telegram.admins
            or self.base.is_payment_admin(self.update.user, self.pass_key)
        ):
            return await self.update.reply(
                f"{self.update.user} is not admin for {self.pass_key}",
                parse_mode=None,
            )
        pass_doc = await self.base.pass_db.find_one(
            {
                "bot_id": self.bot,
                "user_id": target_user_id,
                "pass_key": self.pass_key,
            }
        )
        if pass_doc is None:
            return await self.update.reply(
                f"User {target_user_id} does not have pass {self.pass_key}",
                parse_mode=None,
            )
        pass_data = self.base._pass_doc_to_data(pass_doc) or {}
        uids = [pass_doc["user_id"]]
        if "couple" in pass_data:
            uids.append(pass_data["couple"])
        admin_doc = await self.base.user_db.find_one(
            {
                "bot_id": self.bot,
                "user_id": self.update.user,
            }
        )
        await self.base.update_pass_fields(
            uids,
            self.pass_key,
            set_fields={"proof_admin": self.update.user},
        )
        for uid in uids:
            try:
                upd = await self.base.create_update_from_user(uid)
                upd.set_pass_key(self.pass_key)
                await upd.update.reply(
                    await upd.format_message(
                        "passes-admin-changed",
                        admin=admin_doc,
                    ),
                    parse_mode=ParseMode.HTML,
                )
            except Exception as err:
                logger.error(
                    "Failed to notify user %s about new admin: %s", uid, err, exc_info=1
                )
        await self.update.reply(
            f"Admin for pass {self.pass_key} switched to you for users {uids}",
            parse_mode=None,
        )
    
    async def handle_passes_table_cmd(self):
        self.base.refresh_events_cache()
        if self.update.user in self.base.config.telegram.admins:
            pass_keys_for_this_user = self.base.pass_keys
        else:
            pass_keys_for_this_user = [
                key
                for key in self.base.pass_keys
                if key in self.base.payment_admins
                and self.base.is_payment_admin(self.update.user, key)
            ]
        assert pass_keys_for_this_user, (
            f"{self.update.user} is not admin for any pass key"
        )
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Passes"

        fields = {
            "user_id": {"name": "User ID", "location": "user"},
            "username": {"name": "Username", "location": "user"},
            "first_name": {"name": "First Name", "location": "user"},
            "last_name": {"name": "Last Name", "location": "user"},
            "print_name": {"name": "Print Name", "location": "user"},
            "legal_name": {"name": "Legal Name", "location": "user"},
            "language_code": {"name": "Language code", "location": "user", "length": 2},
            "pass_key": {
                "func": lambda u, p, k: k,
                "name": "Pass Key",
            },
            "state": {"name": "State", "location": "pass"},
            "type": {"name": "Type", "location": "pass"},
            "role": {"name": "Role", "location": "pass"},
            "couple": {"name": "Couple", "location": "pass"},
            "price": {"name": "Price", "location": "pass"},
            "price_per_one": {"name": "Price per one", "location": "pass"},
            "assignment_tier_number": {"name": "Assignment Tier", "location": "pass"},
            "date_created": {"name": "Date Created", "location": "pass"},
            "date_assignment": {"name": "Date Assignment", "location": "pass"},
            "proof_admin": {"name": "Proof Admin", "location": "pass"},
            "proof_received": {"name": "Proof Received", "location": "pass"},
            "proof_accepted": {"name": "Proof Accepted", "location": "pass"},
            "proof_rejected": {"name": "Proof Rejected", "location": "pass"},
            "proof_file": {"name": "Proof File ID", "location": "pass", "length": 5},
            "skip_in_balance_count": {"name": "Skip in Balance Count", "location": "pass"},
            "comment": {"name": "Comment", "location": "pass"},
        }

        bold = Font(bold=True)
        center = Alignment(horizontal="center")
        ws.append([field["name"] for field in fields.values()])
        for cell in ws["1:1"]:
            cell.font = bold
            cell.alignment = center

        for i, field in enumerate(fields.values()):
            field["column_number"] = i + 1
            field["column_letter"] = ws.cell(row=1, column=i + 1).column_letter

        for pass_key in pass_keys_for_this_user:
            pass_docs = await self.base.pass_db.find(
                {
                    "bot_id": self.bot,
                    "pass_key": pass_key,
                }
            ).to_list(None)
            pass_data_by_uid = {
                doc["user_id"]: self.base._pass_doc_to_data(doc) or {}
                for doc in pass_docs
            }
            for pass_doc in pass_docs:
                user = await self.base.user_db.find_one(
                    {"bot_id": self.bot, "user_id": pass_doc.get("user_id")}
                ) or {}
                pass_data = pass_data_by_uid.get(pass_doc.get("user_id"), {})
                couple_pass = pass_data_by_uid.get(pass_data.get("couple"))
                display_price = self.base.get_pass_display_price(pass_data, couple_pass)
                row = []
                for field, info in fields.items():
                    if field == "price":
                        row.append(display_price)
                    elif field == "price_per_one":
                        row.append(pass_data.get("price", ""))
                    elif "func" in info:
                        row.append(info["func"](user, pass_data, pass_key))
                    elif info["location"] == "user":
                        row.append(user.get(field, ""))
                    elif info["location"] == "pass":
                        row.append(pass_data.get(field, ""))
                ws.append(row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.add_sort_condition(fields["date_created"]["column_letter"] + "2:" + fields["date_created"]["column_letter"] + str(ws.max_row))
        for i, field in enumerate(fields.values()):
            if "length" in field:
                ws.column_dimensions[field["column_letter"]].width = field["length"]
            else: # auto width
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in ws[field["column_letter"]]
                )
                ws.column_dimensions[field["column_letter"]].width = max(max_length, 3)

        wb.save("passes.xlsx")
        await self.update.bot.send_document(
            self.update.user,
            open("passes.xlsx", "rb"),
            caption="Passes table",
        )
        import os
        os.remove("passes.xlsx")


class Passes(BasePlugin):
    name = "passes"

    def __init__(self, base_app):
        super().__init__(base_app)
        self.base_app.passes = self
        self.events: Events = self.base_app.events
        self.pass_keys: list[str] = []
        self.payment_admins: dict[str, list[int]] = {}
        self.hidden_payment_admins: dict[str, list[int]] = {}
        if base_app.passes_collection is None:
            raise RuntimeError("passes_collection is not configured")
        self.pass_db: AgnosticCollection = base_app.passes_collection
        self.refresh_events_cache()
        self.user_db: AgnosticCollection = base_app.users_collection
        self._command_checkers = [
            CommandHandler(self.name, self.handle_start),
            CommandHandler("passes_assign", self.handle_passes_assign_cmd),
            CommandHandler("passes_cancel", self.handle_passes_cancel_cmd),
            CommandHandler("passes_tier", self.handle_passes_tier_cmd),
            CommandHandler("passes_switch_to_me", self.handle_passes_switch_to_me_cmd),
            CommandHandler("passes_uncouple", self.handle_passes_uncouple_cmd),
            CommandHandler("passes_table", self.handle_passes_table_cmd),
            CommandHandler("legal_name", self.handle_name_cmd),
            CommandHandler("passport", self.handle_passport_data_cmd),
            CommandHandler("role", self.handle_role_cmd),
        ]
        self._cbq_handler = CallbackQueryHandler(
            self.handle_callback_query, pattern=f"^{self.name}\\|.*"
        )
        self._queue_lock = Lock()
        self._queue_recalc_pending = False
        create_task(self._timeout_processor())

    def refresh_events_cache(self) -> None:
        active_events: list[EventInfo] = list(self.events.active_events(now_msk()))
        self.pass_keys = [event.key for event in active_events]
        self.payment_admins = {
            event.key: list(event.payment_admin) for event in active_events
        }
        self.hidden_payment_admins = {
            event.key: list(event.hidden_payment_admins) for event in active_events
        }

    def all_pass_keys(self) -> list[str]:
        return self.events.all_pass_keys()

    def default_pass_key(self) -> str:
        return self.events.closest_active_pass_key(now_msk())

    def get_event(self, pass_key: str) -> EventInfo | None:
        return self.events.get_event(pass_key)

    def require_event(self, pass_key: str) -> EventInfo:
        event = self.get_event(pass_key)
        if event is None:
            raise KeyError(f"Unknown event for pass key {pass_key}")
        return event

    def get_event_localization_keys(
        self,
        pass_key: str,
        locale: str | None = None,
    ) -> dict[str, str]:
        event = self.get_event(pass_key)
        if event is None:
            return {
                "eventTitleLong": pass_key,
                "eventTitleShort": pass_key,
                "eventCountryEmoji": "",
            }
        return {
            "eventTitleLong": event.title_long_for_locale(locale),
            "eventTitleShort": event.title_short_for_locale(locale),
            "eventCountryEmoji": event.country_emoji,
        }

    def _pass_doc_to_data(self, pass_doc: dict | None) -> dict | None:
        if pass_doc is None:
            return None
        data = dict(pass_doc)
        for key in ["_id", "user_id", "pass_key", "bot_id"]:
            data.pop(key, None)
        return data

    @staticmethod
    def _price_value(value) -> int | float | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return value
        return None

    def get_pass_display_price(
        self,
        pass_data: dict | None,
        couple_pass: dict | None = None,
    ):
        if not isinstance(pass_data, dict):
            return ""
        own_raw = pass_data.get("price")
        own_price = self._price_value(own_raw)
        if own_price is None:
            return own_raw if own_raw is not None else ""
        if "couple" in pass_data and isinstance(couple_pass, dict):
            partner_price = self._price_value(couple_pass.get("price"))
            if partner_price is not None:
                return own_price + partner_price
        return own_price

    @staticmethod
    def split_total_price(total_price: int, user_ids: list[int]) -> dict[int, int]:
        if not user_ids:
            return {}
        sorted_uids = sorted(user_ids)
        base = total_price // len(sorted_uids)
        remainder = total_price - base * len(sorted_uids)
        return {
            uid: base + (1 if i < remainder else 0)
            for i, uid in enumerate(sorted_uids)
        }

    def default_price_per_person(self, pass_key: str) -> int:
        event = self.get_event(pass_key)
        if event is not None and event.pass_types:
            tier_price = event.pass_types[0].price
            if tier_price <= 0:
                raise ValueError(
                    f"invalid tier price for pass key {pass_key}: {tier_price}; "
                    "tier prices must be > 0"
                )
            return tier_price
        event_price = event.price if event is not None else None
        if isinstance(event_price, int):
            if event_price <= 0:
                raise ValueError(
                    f"invalid default event price for pass key {pass_key}: {event_price}; "
                    "automatic pricing must be > 0"
                )
            return event_price
        if pass_key in CURRENT_PRICE:
            fallback_price = CURRENT_PRICE[pass_key]
            if fallback_price <= 0:
                raise ValueError(
                    f"invalid fallback price for pass key {pass_key}: {fallback_price}; "
                    "automatic pricing must be > 0"
                )
            return fallback_price
        raise KeyError(f"no default price configured for pass key {pass_key}")

    def resolve_total_pass_price(
        self,
        pass_key: str,
        user_ids: list[int],
        custom_total_price: int | None = None,
    ) -> int:
        if custom_total_price is not None:
            return custom_total_price
        return self.default_price_per_person(pass_key) * len(user_ids)

    def _event_assignment_rule(self, event: EventInfo) -> str:
        if event.pass_assignment_rule in {"paired", "distributed"}:
            return event.pass_assignment_rule
        return "paired"

    def _event_pass_types(self, event: EventInfo) -> tuple[EventPassType, ...]:
        return event.pass_types

    @staticmethod
    def _safe_tier_index(raw_index: object) -> int | None:
        if isinstance(raw_index, bool):
            return None
        if isinstance(raw_index, int) and raw_index >= 0:
            return raw_index
        return None

    @staticmethod
    def _role_total_from_counts(
        role_counts: dict[str, dict[str, int]],
        role: str,
    ) -> int:
        return role_counts.get(role, {}).get("RA", 0)

    @staticmethod
    def _is_within_balance_tolerance(leader_count: int, follower_count: int) -> bool:
        total = leader_count + follower_count
        if total <= 1:
            return True
        higher_role = max(leader_count, follower_count)
        return higher_role * 100 <= ROLE_BALANCE_TOLERANCE * total

    def _can_assign_with_balance(
        self,
        role_counts: dict[str, dict[str, int]],
        *,
        leader_delta: int = 0,
        follower_delta: int = 0,
    ) -> bool:
        leader_count = self._role_total_from_counts(role_counts, "leader")
        follower_count = self._role_total_from_counts(role_counts, "follower")
        next_leader_count = leader_count + leader_delta
        next_follower_count = follower_count + follower_delta

        if self._is_within_balance_tolerance(leader_count, follower_count):
            return self._is_within_balance_tolerance(
                next_leader_count,
                next_follower_count,
            )

        # When already outside tolerance, allow only non-worsening moves.
        current_diff = abs(leader_count - follower_count)
        next_diff = abs(next_leader_count - next_follower_count)
        if next_diff > current_diff:
            return False
        current_total = leader_count + follower_count
        next_total = next_leader_count + next_follower_count
        if current_total <= 0 or next_total <= 0:
            return True
        current_higher = max(leader_count, follower_count)
        next_higher = max(next_leader_count, next_follower_count)
        return next_higher * current_total <= current_higher * next_total

    @staticmethod
    def _tier_prior_capacity(
        pass_types: tuple[EventPassType, ...],
        tier_index: int,
        assignment_rule: str,
    ) -> int:
        if assignment_rule == "paired":
            return sum(pass_type.amount // 2 for pass_type in pass_types[:tier_index])
        return sum(pass_type.amount for pass_type in pass_types[:tier_index])

    @staticmethod
    def _tier_capacity_for_rule(
        pass_type: EventPassType,
        assignment_rule: str,
    ) -> int:
        if assignment_rule == "paired":
            return pass_type.amount // 2
        return pass_type.amount

    def _tier_implied_usage(
        self,
        *,
        pass_types: tuple[EventPassType, ...],
        assignment_rule: str,
        tier_index: int,
        participants: int,
    ) -> int:
        if participants <= 0:
            return 0
        tier_limit = self._tier_capacity_for_rule(pass_types[tier_index], assignment_rule)
        if tier_limit <= 0:
            return 0
        sold_before = self._tier_prior_capacity(pass_types, tier_index, assignment_rule)
        usage = participants - sold_before
        if usage <= 0:
            return 0
        return min(usage, tier_limit)

    def _effective_tier_usage(
        self,
        *,
        pass_types: tuple[EventPassType, ...],
        assignment_rule: str,
        tier_usage: dict[int, int],
        tier_index: int,
        participants: int,
    ) -> int:
        explicit_usage = int(tier_usage.get(tier_index, 0))
        implied_usage = self._tier_implied_usage(
            pass_types=pass_types,
            assignment_rule=assignment_rule,
            tier_index=tier_index,
            participants=participants,
        )
        return max(explicit_usage, implied_usage)

    def _tier_display_left(
        self,
        *,
        pass_types: tuple[EventPassType, ...],
        assignment_rule: str,
        tier_usage: dict[int, int],
        tier_index: int,
        participants: int,
    ) -> int:
        """Display left count, including unsold capacity carried from prior tiers."""
        tier_limit = self._tier_capacity_for_rule(pass_types[tier_index], assignment_rule)
        if tier_limit <= 0:
            return 0
        tier_used = self._effective_tier_usage(
            pass_types=pass_types,
            assignment_rule=assignment_rule,
            tier_usage=tier_usage,
            tier_index=tier_index,
            participants=participants,
        )
        sold_before = self._tier_prior_capacity(pass_types, tier_index, assignment_rule)
        carryover = max(sold_before - participants, 0)
        return max(tier_limit - tier_used, 0) + carryover

    def _time_floor_tier_index(
        self,
        *,
        pass_types: tuple[EventPassType, ...],
        allow_promo: bool,
    ) -> int | None:
        now = now_msk()
        floor_index: int | None = None
        first_assignable: int | None = None
        for tier_index, pass_type in enumerate(pass_types):
            if not allow_promo and pass_type.promo:
                continue
            if first_assignable is None:
                first_assignable = tier_index
            if pass_type.start <= now:
                floor_index = tier_index
        if floor_index is not None:
            return floor_index
        return first_assignable

    def _is_tier_effective(
        self,
        *,
        pass_types: tuple[EventPassType, ...],
        tier_index: int,
        assignment_rule: str,
        projected_participants: int,
    ) -> bool:
        pass_type = pass_types[tier_index]
        if pass_type.start <= now_msk():
            return True
        if tier_index == 0:
            return False
        sold_before = self._tier_prior_capacity(pass_types, tier_index, assignment_rule)
        return sold_before < projected_participants

    def _pick_tier_for_assignment(
        self,
        *,
        pass_types: tuple[EventPassType, ...],
        assignment_rule: str,
        tier_usage_total: dict[int, int],
        tier_usage_by_role: dict[str, dict[int, int]],
        participants_total: int,
        participants_by_role: dict[str, int],
        role: str | None,
        increment: int,
        allow_promo: bool,
    ) -> int | None:
        start_tier_index = self._time_floor_tier_index(
            pass_types=pass_types,
            allow_promo=allow_promo,
        )
        if start_tier_index is None:
            return None
        for tier_index in range(start_tier_index, len(pass_types)):
            pass_type = pass_types[tier_index]
            if not allow_promo and pass_type.promo:
                continue
            if assignment_rule == "paired":
                if role not in {"leader", "follower"}:
                    return None
                tier_limit = self._tier_capacity_for_rule(pass_type, assignment_rule)
                if tier_limit <= 0:
                    continue
                projected_participants = participants_by_role.get(role, 0) + increment
                if not self._is_tier_effective(
                    pass_types=pass_types,
                    tier_index=tier_index,
                    assignment_rule=assignment_rule,
                    projected_participants=projected_participants,
                ):
                    continue
                tier_used = self._effective_tier_usage(
                    pass_types=pass_types,
                    assignment_rule=assignment_rule,
                    tier_usage=tier_usage_by_role.get(role, {}),
                    tier_index=tier_index,
                    participants=participants_by_role.get(role, 0),
                )
                if tier_used + increment <= tier_limit:
                    return tier_index
                continue

            tier_limit = self._tier_capacity_for_rule(pass_type, assignment_rule)
            if tier_limit <= 0:
                continue
            projected_participants = participants_total + increment
            if not self._is_tier_effective(
                pass_types=pass_types,
                tier_index=tier_index,
                assignment_rule=assignment_rule,
                projected_participants=projected_participants,
            ):
                continue
            tier_used = self._effective_tier_usage(
                pass_types=pass_types,
                assignment_rule=assignment_rule,
                tier_usage=tier_usage_total,
                tier_index=tier_index,
                participants=participants_total,
            )
            if tier_used + increment <= tier_limit:
                return tier_index
        return None

    @staticmethod
    def _role_counts_from_aggregation(
        aggregation: list[dict[str, object]],
    ) -> tuple[dict[str, dict[str, int]], int]:
        role_counts: dict[str, dict[str, int]] = {
            "leader": {},
            "follower": {},
        }
        max_assigned = 0
        for group in aggregation:
            group_id = group.get("_id")
            if not isinstance(group_id, dict):
                continue
            role = group_id.get("role")
            state = group_id.get("state")
            if role not in {"leader", "follower"} or not isinstance(state, str):
                continue
            count = int(group["count"])
            role_counts[role][state] = count
            if state == "assigned" and max_assigned < count:
                max_assigned = count
        role_counts["leader"]["RA"] = (
            role_counts["leader"].get("paid", 0)
            + role_counts["leader"].get("assigned", 0)
        )
        role_counts["follower"]["RA"] = (
            role_counts["follower"].get("paid", 0)
            + role_counts["follower"].get("assigned", 0)
        )
        return role_counts, max_assigned

    async def _collect_queue_stats(self, pass_key: str) -> dict[str, object]:
        aggregation = await self.pass_db.aggregate(
            [
                {
                    "$match": {
                        "bot_id": self.bot.id,
                        "pass_key": pass_key,
                        "$and": [self._balance_counter_match()],
                    },
                },
                {
                    "$group": {
                        "_id": {
                            "state": "$state",
                            "role": "$role",
                        },
                        "count": {"$count": {}},
                    }
                },
            ]
        ).to_list(None)
        couples = {
            group["_id"]: group["count"]
            for group in await self.pass_db.aggregate(
                [
                    {
                        "$match": {
                            "couple": {"$exists": True},
                            "role": "leader",
                            "bot_id": self.bot.id,
                            "pass_key": pass_key,
                            "$and": [self._balance_counter_match()],
                        },
                    },
                    {
                        "$group": {
                            "_id": "$state",
                            "count": {"$count": {}},
                        }
                    },
                ]
            ).to_list(None)
        }
        role_counts, max_assigned = self._role_counts_from_aggregation(aggregation)
        full_aggregation = await self.pass_db.aggregate(
            [
                {
                    "$match": {
                        "bot_id": self.bot.id,
                        "pass_key": pass_key,
                        "state": {"$in": ["waitlist", "assigned", "paid"]},
                    }
                },
                {
                    "$group": {
                        "_id": {
                            "state": "$state",
                            "role": "$role",
                        },
                        "count": {"$count": {}},
                    }
                },
            ]
        ).to_list(None)
        full_role_counts, _ = self._role_counts_from_aggregation(full_aggregation)

        sold_aggregation = await self.pass_db.aggregate(
            [
                {
                    "$match": {
                        "bot_id": self.bot.id,
                        "pass_key": pass_key,
                        "state": {"$in": ["assigned", "paid"]},
                    },
                },
                {
                    "$group": {
                        "_id": {
                            "role": "$role",
                            "pass_type_index": "$pass_type_index",
                        },
                        "count": {"$count": {}},
                    }
                },
            ]
        ).to_list(None)
        participants_total = 0
        participants_by_role: dict[str, int] = {"leader": 0, "follower": 0}
        tier_usage_total: dict[int, int] = {}
        tier_usage_by_role: dict[str, dict[int, int]] = {
            "leader": {},
            "follower": {},
        }
        for group in sold_aggregation:
            group_id = group.get("_id")
            if not isinstance(group_id, dict):
                continue
            tier_index = self._safe_tier_index(group_id.get("pass_type_index"))
            count = int(group["count"])
            participants_total += count
            if tier_index is not None:
                tier_usage_total[tier_index] = tier_usage_total.get(tier_index, 0) + count
            role = group_id.get("role")
            if role in {"leader", "follower"}:
                participants_by_role[role] = participants_by_role.get(role, 0) + count
                if tier_index is not None:
                    role_tiers = tier_usage_by_role[role]
                    role_tiers[tier_index] = role_tiers.get(tier_index, 0) + count

        return {
            "role_counts": role_counts,
            "full_role_counts": full_role_counts,
            "max_single_assigned": max_assigned - couples.get("assigned", 0),
            "participants_total": participants_total,
            "participants_by_role": participants_by_role,
            "tier_usage_total": tier_usage_total,
            "tier_usage_by_role": tier_usage_by_role,
        }

    @staticmethod
    def _extract_tier_stats(
        stats: dict[str, object] | None,
    ) -> tuple[int, dict[str, int], dict[int, int], dict[str, dict[int, int]]]:
        if stats is None:
            return 0, {"leader": 0, "follower": 0}, {}, {"leader": {}, "follower": {}}
        participants_total = int(stats.get("participants_total", 0))
        raw_participants_by_role = stats.get("participants_by_role", {})
        if isinstance(raw_participants_by_role, dict):
            participants_by_role = {
                "leader": int(raw_participants_by_role.get("leader", 0)),
                "follower": int(raw_participants_by_role.get("follower", 0)),
            }
        else:
            participants_by_role = {"leader": 0, "follower": 0}
        raw_tier_usage_total = stats.get("tier_usage_total", {})
        if isinstance(raw_tier_usage_total, dict):
            tier_usage_total = {
                int(k): int(v) for k, v in raw_tier_usage_total.items()
            }
        else:
            tier_usage_total = {}
        raw_tier_usage_by_role = stats.get("tier_usage_by_role", {})
        tier_usage_by_role: dict[str, dict[int, int]] = {
            "leader": {},
            "follower": {},
        }
        if isinstance(raw_tier_usage_by_role, dict):
            for role_key in ("leader", "follower"):
                role_map = raw_tier_usage_by_role.get(role_key, {})
                if isinstance(role_map, dict):
                    tier_usage_by_role[role_key] = {
                        int(k): int(v) for k, v in role_map.items()
                    }
        return (
            participants_total,
            participants_by_role,
            tier_usage_total,
            tier_usage_by_role,
        )

    @staticmethod
    def _format_tier_start(dt) -> str:
        return dt.strftime("%Y-%m-%d %H:%M")

    @staticmethod
    def _format_balance_details(role_counts: dict[str, dict[str, int]]) -> str:
        leader = int(role_counts.get("leader", {}).get("RA", 0))
        follower = int(role_counts.get("follower", {}).get("RA", 0))
        total = leader + follower
        if total <= 0:
            return "L|F=0|0 (L=0.0%, F=0.0%)"
        leader_pct = (leader * 100.0) / total
        follower_pct = (follower * 100.0) / total
        return f"L|F={leader}|{follower} (L={leader_pct:.1f}%, F={follower_pct:.1f}%)"

    async def format_current_tier_status(self, pass_key: str) -> str:
        event = self.require_event(pass_key)
        pass_types = self._event_pass_types(event)
        assignment_rule = self._event_assignment_rule(event)
        if len(pass_types) == 0:
            return f"pass_key={pass_key}\nNo tiers are configured for this event."
        stats = await self._collect_queue_stats(pass_key)
        (
            participants_total,
            participants_by_role,
            tier_usage_total,
            tier_usage_by_role,
        ) = self._extract_tier_stats(stats)
        role_counts = stats.get("role_counts", {})
        if not isinstance(role_counts, dict):
            role_counts = {"leader": {}, "follower": {}}
        full_role_counts = stats.get("full_role_counts", {})
        if not isinstance(full_role_counts, dict):
            full_role_counts = {"leader": {}, "follower": {}}
        assigned_leader = int(role_counts.get("leader", {}).get("assigned", 0))
        assigned_follower = int(role_counts.get("follower", {}).get("assigned", 0))
        waitlist_leader = int(role_counts.get("leader", {}).get("waitlist", 0))
        waitlist_follower = int(role_counts.get("follower", {}).get("waitlist", 0))
        waitlist_total = waitlist_leader + waitlist_follower
        waitlist_role = "none"
        if waitlist_total > 0:
            waitlist_role = self._target_role(role_counts)
        lines = [
            f"pass key={pass_key}",
            f"assignment rule={assignment_rule}",
            f"balance: {self._format_balance_details(role_counts)}",
            "balance full: "
            f"{self._format_balance_details(full_role_counts)}",
            f"assigned state: L|F={assigned_leader}|{assigned_follower}",
            (
                "waitlist: "
                f"number={waitlist_total}, role={waitlist_role}, "
                f"L|F={waitlist_leader}|{waitlist_follower}"
            ),
            (
                "assigned + paid: "
                f"total={participants_total}, "
                f"leader={participants_by_role.get('leader', 0)}, "
                f"follower={participants_by_role.get('follower', 0)}"
            ),
        ]
        if assignment_rule == "distributed":
            tier_index = self._pick_tier_for_assignment(
                pass_types=pass_types,
                assignment_rule=assignment_rule,
                tier_usage_total=tier_usage_total,
                tier_usage_by_role=tier_usage_by_role,
                participants_total=participants_total,
                participants_by_role=participants_by_role,
                role=None,
                increment=1,
                allow_promo=True,
            )
            if tier_index is None:
                lines.append("current tier: none (no assignable tiers left)")
                return "\n".join(lines)
            tier_info = pass_types[tier_index]
            tier_limit = self._tier_capacity_for_rule(tier_info, assignment_rule)
            tier_left = self._tier_display_left(
                pass_types=pass_types,
                assignment_rule=assignment_rule,
                tier_usage=tier_usage_total,
                tier_index=tier_index,
                participants=participants_total,
            )
            lines.append(
                "current tier: "
                f"{tier_index + 1} "
                f"(price={tier_info.price}, left/total={tier_left}/{tier_limit}, "
                f"promo={tier_info.promo}, start={self._format_tier_start(tier_info.start)})"
            )
            return "\n".join(lines)

        for role in ("leader", "follower"):
            tier_index = self._pick_tier_for_assignment(
                pass_types=pass_types,
                assignment_rule=assignment_rule,
                tier_usage_total=tier_usage_total,
                tier_usage_by_role=tier_usage_by_role,
                participants_total=participants_total,
                participants_by_role=participants_by_role,
                role=role,
                increment=1,
                allow_promo=True,
            )
            if tier_index is None:
                lines.append(f"{role}: current tier: none (no assignable tiers left)")
                continue
            tier_info = pass_types[tier_index]
            tier_limit = self._tier_capacity_for_rule(tier_info, assignment_rule)
            tier_left = self._tier_display_left(
                pass_types=pass_types,
                assignment_rule=assignment_rule,
                tier_usage=tier_usage_by_role.get(role, {}),
                tier_index=tier_index,
                participants=participants_by_role.get(role, 0),
            )
            lines.append(
                f"{role}: current tier: {tier_index + 1} "
                f"(price={tier_info.price}, left/total={tier_left}/{tier_limit}, "
                f"promo={tier_info.promo}, start={self._format_tier_start(tier_info.start)})"
            )
        return "\n".join(lines)

    async def resolve_candidate_tier_prices(
        self,
        pass_key: str,
        user_id: int,
        pass_data: dict,
    ) -> tuple[dict[int, int], dict[int, int]] | None:
        event = self.require_event(pass_key)
        pass_types = self._event_pass_types(event)
        assignment_rule = self._event_assignment_rule(event)
        if len(pass_types) == 0:
            return None

        user_roles: dict[int, str] = {}
        role = pass_data.get("role")
        if role not in {"leader", "follower"}:
            return None
        user_roles[user_id] = role

        is_couple = "couple" in pass_data
        if is_couple:
            couple_user_id = pass_data.get("couple")
            if not isinstance(couple_user_id, int):
                return None
            couple_pass_doc = await self.pass_db.find_one(
                {
                    "bot_id": self.bot.id,
                    "pass_key": pass_key,
                    "user_id": couple_user_id,
                    "state": "waitlist",
                    "couple": user_id,
                }
            )
            if not isinstance(couple_pass_doc, dict):
                return None
            couple_pass_info = self._pass_doc_to_data(couple_pass_doc) or {}
            couple_role = couple_pass_info.get("role")
            if couple_role not in {"leader", "follower"}:
                return None
            user_roles[couple_user_id] = couple_role

        (
            participants_total,
            participants_by_role,
            tier_usage_total,
            tier_usage_by_role,
        ) = self._extract_tier_stats(await self._collect_queue_stats(pass_key))

        price_by_user: dict[int, int] = {}
        pass_type_index_by_user: dict[int, int] = {}
        if assignment_rule == "distributed":
            tier_index = self._pick_tier_for_assignment(
                pass_types=pass_types,
                assignment_rule=assignment_rule,
                tier_usage_total=tier_usage_total,
                tier_usage_by_role=tier_usage_by_role,
                participants_total=participants_total,
                participants_by_role=participants_by_role,
                role=None,
                increment=len(user_roles),
                allow_promo=not is_couple,
            )
            if tier_index is None:
                return None
            tier_price = pass_types[tier_index].price
            for uid in user_roles:
                price_by_user[uid] = tier_price
                pass_type_index_by_user[uid] = tier_index
            return price_by_user, pass_type_index_by_user

        allow_promo = not is_couple
        for uid, user_role in sorted(user_roles.items()):
            tier_index = self._pick_tier_for_assignment(
                pass_types=pass_types,
                assignment_rule=assignment_rule,
                tier_usage_total=tier_usage_total,
                tier_usage_by_role=tier_usage_by_role,
                participants_total=participants_total,
                participants_by_role=participants_by_role,
                role=user_role,
                increment=1,
                allow_promo=allow_promo,
            )
            if tier_index is None:
                return None
            price_by_user[uid] = pass_types[tier_index].price
            pass_type_index_by_user[uid] = tier_index
        return price_by_user, pass_type_index_by_user

    def _target_role(self, role_counts: dict[str, dict[str, int]]) -> str:
        leader_ra = self._role_total_from_counts(role_counts, "leader")
        follower_ra = self._role_total_from_counts(role_counts, "follower")
        total_ra = leader_ra + follower_ra

        if total_ra > 0 and not self._is_within_balance_tolerance(leader_ra, follower_ra):
            return "follower" if leader_ra > follower_ra else "leader"

        leader_waitlist = role_counts.get("leader", {}).get("waitlist", 0)
        follower_waitlist = role_counts.get("follower", {}).get("waitlist", 0)
        if leader_waitlist == follower_waitlist:
            return "leader" if leader_ra <= follower_ra else "follower"
        return "leader" if leader_waitlist > follower_waitlist else "follower"

    @staticmethod
    def _balance_counter_match() -> dict[str, object]:
        return {
            "$or": [
                {"skip_in_balance_count": False},
                {
                    "$and": [
                        {"skip_in_balance_count": {"$ne": True}},
                        {"price": {"$ne": 0}},
                    ]
                },
            ]
        }

    async def get_pass_for_user(self, user_id: int, pass_key: str) -> dict | None:
        doc = await self.pass_db.find_one(
            {
                "bot_id": self.bot.id,
                "user_id": user_id,
                "pass_key": pass_key,
            }
        )
        if doc is None or "pass_key" not in doc:
            return None
        return self._pass_doc_to_data(doc)

    async def save_pass_data(
        self,
        user_id: int,
        pass_key: str,
        pass_data: dict,
    ) -> None:
        doc = dict(pass_data)
        doc["user_id"] = user_id
        doc["pass_key"] = pass_key
        doc["bot_id"] = self.bot.id
        await self.pass_db.update_one(
            {"bot_id": self.bot.id, "user_id": user_id, "pass_key": pass_key},
            {"$set": doc},
            upsert=True,
        )

    async def get_user_with_pass(self, user_id: int, pass_key: str) -> tuple[dict | None, dict | None]:
        user = await self.user_db.find_one(
            {
                "bot_id": self.bot.id,
                "user_id": user_id,
            }
        )
        if user is None:
            return None, None
        pass_data = await self.get_pass_for_user(user_id, pass_key)
        if pass_data is None and pass_key in user:
            pass_data = user.get(pass_key)
            if isinstance(pass_data, dict):
                logger.warning("Found embedded passes data for user %s pass %s, migrating", user_id, pass_key)
                await self.save_pass_data(user_id, pass_key, pass_data)
        return user, pass_data

    async def update_pass_fields(
        self,
        user_ids: list[int],
        pass_key: str,
        *,
        set_fields: dict | None = None,
        unset_fields: list[str] | None = None,
    ) -> None:
        if not user_ids:
            return
        update: dict = {}
        if set_fields:
            update["$set"] = dict(set_fields)
        if unset_fields:
            update.setdefault("$unset", {})
            for field in unset_fields:
                update["$unset"][field] = ""
        if update:
            for uid in user_ids:
                await self.pass_db.update_one(
                    {
                        "bot_id": self.bot.id,
                        "pass_key": pass_key,
                        "user_id": uid,
                    },
                    update,
                    upsert=True,
                )

    async def delete_passes(self, user_ids: list[int], pass_key: str) -> None:
        if not user_ids:
            return
        await self.pass_db.delete_many(
            {
                "bot_id": self.bot.id,
                "pass_key": pass_key,
                "user_id": {"$in": user_ids},
            }
        )

    def get_all_payment_admins(self, pass_key: str) -> list[int]:
        admins = self.payment_admins.get(pass_key, [])
        hidden = self.hidden_payment_admins.get(pass_key, [])
        # dict.fromkeys preserves order while removing duplicates
        return list(dict.fromkeys(admins + hidden))

    def is_payment_admin(self, user_id: int, pass_key: str) -> bool:
        return user_id in self.get_all_payment_admins(pass_key)

    def pick_payment_admin(self, pass_key: str) -> int | None:
        visible = self.payment_admins.get(pass_key, [])
        if len(visible) > 0:
            return choice(visible)
        hidden = self.hidden_payment_admins.get(pass_key, [])
        if len(hidden) > 0:
            logger.warning(
                "No visible payment_admins for %s, falling back to hidden_payment_admins",
                pass_key,
            )
            return choice(hidden)
        raise Exception("No payment admins configured for pass %s", pass_key)

    async def get_all_passes(
        self,
        pass_key: str | None = None,
        with_unpaid: bool = False,
        with_waitlist: bool = False,
    ) -> list[dict]:
        pass_key = pass_key or self.default_pass_key()
        states = ["paid"]
        if with_unpaid:
            states.append("assigned")
        query = {
            "bot_id": self.bot.id,
            "pass_key": pass_key,
        }
        if not with_waitlist:
            query["state"] = {"$in": states}
        return await self.pass_db.find(query).to_list(None)

    async def _timeout_processor(self) -> None:
        bot_started: Event = self.base_app.bot_started
        await bot_started.wait()
        self.refresh_events_cache()
        logger.info("timeout processor started")
        migration_result = await self.pass_db.update_many(
            {
                "bot_id": self.bot.id,
                "state": "payed",
            },
            {"$set": {"state": "paid"}},
        )
        if migration_result.modified_count:
            logger.info(
                "Migrated %s passes from payed to paid",
                migration_result.modified_count,
            )
        # await self.migrate_embedded_passes()
        for pass_key in self.pass_keys:
            async for pass_doc in self.pass_db.find(
                {
                    "bot_id": self.bot.id,
                    "pass_key": pass_key,
                    "proof_admin": {"$exists": False},
                }
            ):
                try:
                    new_admin = self.pick_payment_admin(pass_key)
                    await self.update_pass_fields(
                        [pass_doc["user_id"]],
                        pass_key,
                        set_fields={"proof_admin": new_admin},
                    )
                except Exception as e:
                    logger.error(
                        "Exception in Passes._timeout_processor setting initial proof_admin "+
                        f"for user {pass_doc['user_id']}, pass {pass_key}: {e}",
                        exc_info=1,
                    )
        processed_waiting: set[int] = set()
        for pass_key in self.pass_keys:
            all_payment_admins = self.get_all_payment_admins(pass_key)
            async for pass_doc in self.pass_db.find(
                {
                    "bot_id": self.bot.id,
                    "pass_key": pass_key,
                    "state": "waitlist",
                    "$or": [
                        {"proof_admin": {"$nin": all_payment_admins}},
                        {"proof_admin": {"$exists": False}},
                    ],
                }
            ):
                try:
                    if pass_doc["user_id"] in processed_waiting:
                        continue
                    uids = [pass_doc["user_id"]]
                    if "couple" in pass_doc:
                        uids.append(pass_doc["couple"])
                    processed_waiting.update(uids)
                    new_admin = self.pick_payment_admin(pass_key)
                    if new_admin is None:
                        logger.error(
                            "Cannot reassign proof_admin for waitlist user %s, pass %s: no admins configured",
                            pass_doc["user_id"],
                            pass_key,
                        )
                        continue
                    await self.update_pass_fields(
                        uids,
                        pass_key,
                        set_fields={"proof_admin": new_admin},
                    )
                    admin_doc = await self.user_db.find_one(
                        {
                            "bot_id": self.bot.id,
                            "user_id": new_admin,
                        }
                    )
                    for uid in uids:
                        try:
                            upd = await self.create_update_from_user(uid)
                            upd.set_pass_key(pass_key)
                            await upd.update.reply(
                                await upd.format_message(
                                    "passes-admin-changed",
                                    admin=admin_doc,
                                ),
                                parse_mode=ParseMode.HTML,
                            )
                        except Exception as e:
                            logger.error(
                                "Failed to notify user %s about new admin for %s: %s",
                                uid, pass_key, e, exc_info=1,
                            )
                except Exception as e:
                    logger.error(
                        "Exception in Passes._timeout_processor reassigning proof_admin "+
                        f"for waitlist user {pass_doc['user_id']}, pass {pass_key}: {e}",
                        exc_info=1,
                    )

        for pass_key in self.pass_keys:
            event = self.get_event(pass_key)
            if event is None or not event.require_passport:
                continue
            async for pass_doc in self.pass_db.find(
                {
                    "bot_id": self.bot.id,
                    "pass_key": pass_key,
                    "state": {"$in": ["assigned", "paid"]},
                }
            ):
                try:
                    user = await self.user_db.find_one(
                        {
                            "bot_id": self.bot.id,
                            "user_id": pass_doc["user_id"],
                            "passport_number": {"$exists": False},
                            "notified_passport_data_required": {"$exists": False},
                        }
                    )
                    if user is None:
                        continue
                    upd = await self.create_update_from_user(pass_doc["user_id"])
                    upd.set_pass_key(pass_key)
                    await upd.require_passport_data()
                    await self.user_db.update_one(
                        {
                            "bot_id": self.bot.id,
                            "user_id": pass_doc["user_id"],
                            "passport_number": {"$exists": False},
                            "notified_passport_data_required": {"$exists": False},
                        },
                        {
                            "$set": {
                                "notified_passport_data_required": True
                            }
                        }
                    )
                except Exception as e:
                    logger.error(
                        "Exception in Passes._timeout_processor passport notification "+
                        f"for user {pass_doc['user_id']}: {e}",
                        exc_info=1,
                    )

        await self.recalculate_queues()
        while True:
            self.refresh_events_cache()
            if len(self.pass_keys) == 0:
                await sleep(TIMEOUT_PROCESSOR_TICK)
                continue
            for pass_key in self.pass_keys:
                try:
                    await sleep(TIMEOUT_PROCESSOR_TICK)
                    async for pass_doc in self.pass_db.find(
                        {
                            "bot_id": self.bot.id,
                            "pass_key": pass_key,
                            "state": "assigned",
                            "notified_deadline_close": {
                                "$lt": now_msk()
                                - PAYMENT_TIMEOUT
                                + PAYMENT_TIMEOUT_NOTIFY,
                            },
                        }
                    ):
                        upd = await self.create_update_from_user(pass_doc["user_id"])
                        await upd.cancel_due_deadline(pass_key)
                    async for pass_doc in self.pass_db.find(
                        {
                            "bot_id": self.bot.id,
                            "pass_key": pass_key,
                            "state": "waiting-for-couple",
                            "date_created": {
                                "$lt": now_msk() - INVITATION_TIMEOUT,
                            },
                        }
                    ):
                        upd = await self.create_update_from_user(pass_doc["user_id"])
                        await upd.decline_due_deadline(pass_key)
                        if "couple" in pass_doc:
                            upd_c = await self.create_update_from_user(
                                pass_doc["couple"]
                            )
                            await upd_c.decline_invitation_due_deadline(pass_key)
                    async for pass_doc in self.pass_db.find(
                        {
                            "bot_id": self.bot.id,
                            "pass_key": pass_key,
                            "state": "assigned",
                            "date_assignment": {
                                "$lt": now_msk() - PAYMENT_TIMEOUT_NOTIFY,
                            },
                            "notified_deadline_close": {"$exists": False},
                        }
                    ):
                        upd = await self.create_update_from_user(pass_doc["user_id"])
                        upd.set_pass_key(pass_key)
                        await upd.notify_deadline_close()
                    async for pass_doc in self.pass_db.find(
                        {
                            "bot_id": self.bot.id,
                            "pass_key": pass_key,
                            "state": "assigned",
                            "notified_deadline_close": {
                                "$lt": now_msk()
                                - PAYMENT_TIMEOUT_NOTIFY2
                                + PAYMENT_TIMEOUT_NOTIFY,
                            },
                            "notified_deadline_close2": {"$exists": False},
                        }
                    ):
                        upd = await self.create_update_from_user(pass_doc["user_id"])
                        upd.set_pass_key(pass_key)
                        await upd.notify_deadline_close("2")
                except Exception as e:
                    logger.error(
                        "Exception in Passes._timeout_processor %s", e, exc_info=1
                    )

    async def recalculate_queues(self) -> None:
        if self._queue_lock.locked():
            self._queue_recalc_pending = True
            logger.debug(
                "recalculate_queues coalesced: queued rerun after current recalculation"
            )
            return
        async with self._queue_lock:
            while True:
                self._queue_recalc_pending = False
                self.refresh_events_cache()
                for key in self.pass_keys:
                    await self.recalculate_queues_pk(key)
                if not self._queue_recalc_pending:
                    break
                logger.debug(
                    "recalculate_queues rerunning due to coalesced recalculation request"
                )

    async def recalculate_queues_pk(self, pass_key: str) -> None:
        event = self.require_event(pass_key)
        try:
            while True:
                stats = await self._collect_queue_stats(pass_key)
                role_counts = stats.get("role_counts", {})
                if not isinstance(role_counts, dict):
                    break
                max_single_assigned = int(stats.get("max_single_assigned", 0))
                leader_ra = self._role_total_from_counts(role_counts, "leader")
                follower_ra = self._role_total_from_counts(role_counts, "follower")
                if (
                    max_single_assigned > MAX_CONCURRENT_ASSIGNMENTS
                    and self._is_within_balance_tolerance(leader_ra, follower_ra)
                ):
                    break

                target_group = self._target_role(role_counts)
                success = await self.assign_pass(target_group, pass_key, stats)
                if success:
                    continue

                other_group = "follower" if target_group == "leader" else "leader"
                success = await self.assign_pass(other_group, pass_key, stats)
                if success:
                    continue

                success = await self.assign_pass("couple", pass_key, stats)
                if success:
                    continue
                break

            have_unnotified = True
            while have_unnotified:
                selected = await self.pass_db.find(
                    {
                        "bot_id": self.bot.id,
                        "pass_key": pass_key,
                        "state": "waitlist",
                        "no_more_passes_notification_sent": {
                            "$exists": False
                        },
                    }
                ).to_list(100)
                if len(selected) == 0:
                    logger.info("no unnotified candidates in the waiting list")
                    break
                for pass_doc in selected:
                    upd = await self.create_update_from_user(pass_doc["user_id"])
                    upd.set_pass_key(pass_key)
                    await upd.notify_no_more_passes(pass_key)
        except Exception as e:
            logger.error(f"Exception in recalculate_queues: {e}", exc_info=1)

        try:
            async for pass_doc in self.pass_db.find(
                {
                    "bot_id": self.bot.id,
                    "pass_key": pass_key,
                    "sent_to_hype_thread": {"$exists": False},
                }
            ):
                sent_ts = now_msk()
                await self.update_pass_fields(
                    [pass_doc["user_id"]],
                    pass_key,
                    set_fields={"sent_to_hype_thread": sent_ts},
                )
                user = await self.user_db.find_one(
                    {"bot_id": self.bot.id, "user_id": pass_doc["user_id"]}
                ) or {}
                if event.thread_channel != "":
                    try:
                        ch = event.thread_channel
                        if isinstance(ch, str):
                            ch = "@" + ch
                        logger.debug(f"chat id: {ch}, type {type(ch)}")
                        args = {
                            "name": client_user_name(user),
                            "role": pass_doc.get("role", ""),
                            "passKey": pass_key,
                        }
                        await self.bot.send_message(
                            chat_id=ch,
                            message_thread_id=event.thread_id,
                            text=self.base_app.localization(
                                "passes-announce-user-registered",
                                args=args,
                                locale=event.thread_locale,
                            ),
                            parse_mode=ParseMode.HTML,
                        )
                    except Exception as e:
                        logger.error(
                            f"Exception in recalculate_queues: {e}", exc_info=1
                        )
        except Exception as e:
            logger.error(f"Exception in recalculate_queues: {e}", exc_info=1)

    async def assign_pass(
        self,
        role: str,
        pass_key: str,
        stats: dict[str, object] | None = None,
    ) -> bool:
        event = self.require_event(pass_key)
        pass_types = self._event_pass_types(event)
        assignment_rule = self._event_assignment_rule(event)
        if len(pass_types) == 0:
            return False

        match = {
            "bot_id": self.bot.id,
            "pass_key": pass_key,
            "state": "waitlist",
        }
        if role == "couple":
            match["role"] = "leader"
            match["couple"] = {"$exists": True}
        else:
            match["role"] = role
        pipeline = [
            {"$match": match},
            {"$sort": {"date_created": 1, "user_id": 1}},
        ]
        if stats is None:
            stats = await self._collect_queue_stats(pass_key)
        role_counts = stats.get("role_counts", {})
        if not isinstance(role_counts, dict):
            role_counts = {"leader": {}, "follower": {}}
        (
            participants_total,
            participants_by_role,
            tier_usage_total,
            tier_usage_by_role,
        ) = self._extract_tier_stats(stats)

        scanned = 0
        async for candidate in self.pass_db.aggregate(pipeline):
            scanned += 1
            pass_info = self._pass_doc_to_data(candidate) or {}
            is_couple = "couple" in pass_info
            candidate_user_id = candidate.get("user_id")
            if not isinstance(candidate_user_id, int):
                continue

            price_by_user: dict[int, int] = {}
            pass_type_index_by_user: dict[int, int] = {}
            user_roles: dict[int, str] = {}

            candidate_role = pass_info.get("role")
            if candidate_role not in {"leader", "follower"}:
                continue
            user_roles[candidate_user_id] = candidate_role

            if is_couple:
                couple_user_id = pass_info.get("couple")
                if not isinstance(couple_user_id, int):
                    continue
                couple_pass_doc = await self.pass_db.find_one(
                    {
                        "bot_id": self.bot.id,
                        "pass_key": pass_key,
                        "user_id": couple_user_id,
                        "state": "waitlist",
                        "couple": candidate_user_id,
                    }
                )
                if couple_pass_doc is None:
                    continue
                couple_pass_info = self._pass_doc_to_data(couple_pass_doc) or {}
                couple_role = couple_pass_info.get("role")
                if couple_role not in {"leader", "follower"}:
                    continue
                user_roles[couple_user_id] = couple_role

            leader_delta = 1 if "leader" in user_roles.values() else 0
            follower_delta = 1 if "follower" in user_roles.values() else 0
            if not self._can_assign_with_balance(
                role_counts,
                leader_delta=leader_delta,
                follower_delta=follower_delta,
            ):
                continue

            if assignment_rule == "distributed":
                tier_index = self._pick_tier_for_assignment(
                    pass_types=pass_types,
                    assignment_rule=assignment_rule,
                    tier_usage_total=tier_usage_total,
                    tier_usage_by_role=tier_usage_by_role,
                    participants_total=participants_total,
                    participants_by_role=participants_by_role,
                    role=None,
                    increment=len(user_roles),
                    allow_promo=not is_couple,
                )
                if tier_index is None:
                    continue
                tier_price = pass_types[tier_index].price
                for uid in user_roles:
                    price_by_user[uid] = tier_price
                    pass_type_index_by_user[uid] = tier_index
            else:
                allow_promo = not is_couple
                for uid, user_role in user_roles.items():
                    tier_index = self._pick_tier_for_assignment(
                        pass_types=pass_types,
                        assignment_rule=assignment_rule,
                        tier_usage_total=tier_usage_total,
                        tier_usage_by_role=tier_usage_by_role,
                        participants_total=participants_total,
                        participants_by_role=participants_by_role,
                        role=user_role,
                        increment=1,
                        allow_promo=allow_promo,
                    )
                    if tier_index is None:
                        price_by_user = {}
                        pass_type_index_by_user = {}
                        break
                    price_by_user[uid] = pass_types[tier_index].price
                    pass_type_index_by_user[uid] = tier_index
                if not price_by_user:
                    continue

            upd = await self.create_update_from_user(candidate["user_id"])
            assigned = await upd.assign_pass(
                pass_key,
                price_by_user=price_by_user,
                pass_type_index_by_user=pass_type_index_by_user,
            )
            if assigned:
                return True
            if QUEUE_LOOKAHEAD and scanned >= QUEUE_LOOKAHEAD:
                break
        return False

    async def create_update_from_user(self, user: int) -> PassUpdate:
        upd = TGState(user, self.base_app)
        await upd.get_state()
        return PassUpdate(self, upd)

    def test_message(self, message: Update, state, web_app_data):
        for checker in self._command_checkers:
            if checker.check_update(message):
                return PRIORITY_BASIC, checker.callback
        return PRIORITY_NOT_ACCEPTING, None

    def test_callback_query(self, query: Update, state):
        if self._cbq_handler.check_update(query):
            return PRIORITY_BASIC, self.handle_callback_query
        return PRIORITY_NOT_ACCEPTING, None

    def create_update(self, update) -> PassUpdate:
        return PassUpdate(self, update)
    
    async def handle_passport_data_cmd(self, update: TGState):
        return await self.create_update(update).handle_passport_data_cmd()

    async def handle_callback_query(self, updater):
        return await self.create_update(updater).handle_callback_query()

    async def handle_start(self, update: TGState):
        return await self.create_update(update).handle_start()

    async def handle_name_cmd(self, update: TGState):
        return await self.create_update(update).handle_cq_name()

    async def handle_passes_assign_cmd(self, update: TGState):
        return await self.create_update(update).handle_passes_assign()

    async def handle_passes_cancel_cmd(self, update: TGState):
        return await self.create_update(update).handle_passes_cancel()

    async def handle_passes_tier_cmd(self, update: TGState):
        return await self.create_update(update).handle_passes_tier()

    async def handle_passes_switch_to_me_cmd(self, update: TGState):
        return await self.create_update(update).handle_passes_switch_to_me_cmd()
    
    async def handle_passes_uncouple_cmd(self, update: TGState):
        return await self.create_update(update).handle_passes_uncouple()

    async def handle_role_cmd(self, update: TGState):
        return await self.create_update(update).handle_role_cmd()

    async def handle_legal_name_input(self, update, data):
        return await self.create_update(update).handle_legal_name_input(data)

    async def handle_legal_name_timeout(self, update, data):
        return await self.create_update(update).handle_legal_name_timeout(data)

    async def handle_passport_input(self, update, data):
        return await self.create_update(update).handle_passport_input(data)

    async def handle_passport_timeout(self, update, data):
        return await self.create_update(update).handle_passport_timeout(data)

    async def handle_payment_proof_input(self, update, data):
        return await self.create_update(update).handle_payment_proof_input(data)

    async def handle_couple_input(self, update, data):
        return await self.create_update(update).handle_couple_input(data)

    async def handle_couple_timeout(self, update, data):
        return await self.create_update(update).handle_couple_timeout(data)

    async def handle_payment_proof_timeout(self, update, data):
        return await self.create_update(update).handle_payment_proof_timeout(data)
    
    async def handle_passes_table_cmd(self, update: TGState):
        return await self.create_update(update).handle_passes_table_cmd()
